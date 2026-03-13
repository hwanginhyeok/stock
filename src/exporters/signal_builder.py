"""Integrated Signal sheet builder — market trend, stock analysis, sigma bands."""

from __future__ import annotations

from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.exporters.base import (
    DARK_BLUE,
    GRAY,
    GREEN,
    HDR_FILL,
    HDR_FONT,
    LABEL_FILL,
    LABEL_FONT,
    LIGHT_BLUE,
    LIGHT_GRAY,
    MED_BLUE,
    ORANGE,
    PURPLE,
    RED,
    SECTION_FILL,
    SECTION_FONT,
    THIN_BORDER,
    TITLE_FONT,
    VALUE_FONT,
    format_market_cap,
    pct_color_font,
    score_fill,
    style_cell,
    write_label_value,
    write_section_header,
)

# Section-level fill for sigma bands header
_PURPLE_FILL = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
_IV_FONT = Font(bold=True, color="0070C0", size=10)
_HV_FONT = Font(bold=True, color=GRAY, size=10)

# Pattern classification color mapping
_PATTERN_STYLES: dict[str, tuple[str, str]] = {
    # pattern_name: (font_color, fill_color)
    "Breakout": (MED_BLUE, "D6E4F0"),     # blue font, light blue bg
    "Pullback": (GREEN, "C6EFCE"),          # green font, light green bg
    "Consolidation": (ORANGE, "FFF2CC"),    # orange font, light yellow bg
}

# Trend regime color mapping — font color + bold
_TREND_STYLES: dict[str, tuple[str, bool]] = {
    "Strong Uptrend": (GREEN, True),
    "Uptrend": (GREEN, False),
    "Sideways": (GRAY, False),
    "Downtrend": (RED, False),
    "Strong Downtrend": (RED, True),
    "Transition": (ORANGE, False),
}

# Trend regime background fills — pastel colors for ticker cells & legend
_TREND_FILL_COLORS: dict[str, str] = {
    "Strong Uptrend": "C6EFCE",
    "Uptrend": "E2EFDA",
    "Sideways": "F2F2F2",
    "Downtrend": "FBE5D6",
    "Strong Downtrend": "F4CCCC",
    "Transition": "FFF2CC",
}

# Legend display order + Korean labels
_TREND_LEGEND: list[tuple[str, str]] = [
    ("Strong Uptrend", "강한 상승"),
    ("Uptrend", "상승"),
    ("Sideways", "횡보"),
    ("Downtrend", "하락"),
    ("Strong Downtrend", "강한 하락"),
    ("Transition", "전환"),
]


def _trend_font(regime: str) -> Font:
    """Return a styled Font for a trend regime label.

    Args:
        regime: Trend regime string (e.g. "Strong Uptrend").

    Returns:
        Font with appropriate color and boldness.
    """
    color, bold = _TREND_STYLES.get(regime, (GRAY, False))
    return Font(color=color, bold=bold, size=10)


def _trend_fill(regime: str) -> PatternFill:
    """Return a pastel PatternFill for a trend regime.

    Args:
        regime: Trend regime string.

    Returns:
        PatternFill with pastel background.
    """
    hex_color = _TREND_FILL_COLORS.get(regime, "FFFFFF")
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def build_signal_sheet(
    wb: Workbook,
    market_data: dict[str, Any],
    stock_data: list[dict[str, Any]],
    sigma_data: list[dict[str, Any]],
) -> None:
    """Build the Integrated_Signal sheet with 3 sections.

    Args:
        wb: Target Workbook.
        market_data: Dict with keys:
            - index_trend: list of index trend dicts
            - vix: dict with current, level
            - fear_greed: dict from compute_fear_greed()
            - diagnosis: dict from market_diagnosis()
            - putcall: optional dict with ratio, sentiment
            - cnn_fg: optional dict with score, level
        stock_data: List of stock dicts with keys:
            ticker, name, price, change_1d, change_1w, market_cap,
            rsi, rsi_state, macd_signal, supertrend_dir, adx, adx_trend, tech_score
        sigma_data: List of sigma analysis dicts from analyze_ticker().
    """
    ws = wb.create_sheet(title="Integrated_Signal")

    # Title
    ws.merge_cells("A1:O1")
    title_cell = ws.cell(row=1, column=1, value="Integrated Signal Report")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    # Date subtitle
    ws.merge_cells("A2:O2")
    date_cell = ws.cell(
        row=2, column=1,
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    )
    date_cell.font = Font(size=10, color=GRAY)
    date_cell.alignment = Alignment(vertical="center")

    row = 4

    # ===== Section 1: Market Trend & Sentiment =====
    row = _build_market_section(ws, row, market_data)
    row += 2

    # ===== Section 2: Stock Trend & Technical =====
    row = _build_stock_section(ws, row, stock_data)
    row += 2

    # ===== Section 3: Sigma Bands =====
    row = _build_sigma_section(ws, row, sigma_data)

    ws.freeze_panes = "A4"
    ws.sheet_properties.tabColor = PURPLE

    # ===== Guide Sheet =====
    _build_guide_sheet(wb)


# ─── Section 1: Market Trend & Sentiment ──────────────────────


def _build_market_section(
    ws: Any,
    start_row: int,
    market_data: dict[str, Any],
) -> int:
    """Build Section 1: market indices + sentiment indicators + diagnosis.

    Args:
        ws: Worksheet.
        start_row: Starting row.
        market_data: Market data dict.

    Returns:
        Next available row.
    """
    row = start_row

    # Section header
    row = write_section_header(ws, row, "Section 1: Market Trend & Sentiment", cols=10)

    # --- Index table ---
    idx_headers = [
        ("Index", 14), ("Current", 12), ("1D%", 8), ("1W%", 8),
        ("Supertrend", 12), ("ADX", 8), ("ADX Trend", 11),
        ("RSI", 8), ("RSI State", 11), ("Trend", 14),
    ]
    for col_idx, (header, width) in enumerate(idx_headers, 1):
        style_cell(ws, row, col_idx, header, font=HDR_FONT, fill=HDR_FILL,
                   alignment=Alignment(horizontal="center"))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(
            ws.column_dimensions[col_letter].width or 0, width,
        )
    row += 1

    index_trend = market_data.get("index_trend", [])
    for t in index_trend:
        style_cell(ws, row, 1, t["name"], font=Font(bold=True, size=10))
        style_cell(ws, row, 2, t.get("current", 0), num_fmt="#,##0.00",
                   alignment=Alignment(horizontal="right"))

        # 1D%
        chg_1d = t.get("change_1d")
        if chg_1d is not None:
            style_cell(ws, row, 3, chg_1d / 100, font=pct_color_font(chg_1d),
                       num_fmt="+0.00%;-0.00%",
                       alignment=Alignment(horizontal="center"))
        else:
            style_cell(ws, row, 3, "N/A", alignment=Alignment(horizontal="center"))

        # 1W%
        chg_1w = t.get("change_1w")
        if chg_1w is not None:
            style_cell(ws, row, 4, chg_1w / 100, font=pct_color_font(chg_1w),
                       num_fmt="+0.00%;-0.00%",
                       alignment=Alignment(horizontal="center"))
        else:
            style_cell(ws, row, 4, "N/A", alignment=Alignment(horizontal="center"))

        # Supertrend
        st_dir = t.get("supertrend_direction", "N/A")
        st_color = GREEN if st_dir == "Bullish" else (RED if st_dir == "Bearish" else GRAY)
        style_cell(ws, row, 5, st_dir,
                   font=Font(color=st_color, bold=True, size=10),
                   alignment=Alignment(horizontal="center"))

        # ADX
        style_cell(ws, row, 6, t.get("adx", "N/A"),
                   num_fmt="0.0" if t.get("adx") is not None else None,
                   alignment=Alignment(horizontal="center"))

        # ADX Trend
        adx_trend = t.get("adx_trend", "N/A")
        adx_color = GREEN if adx_trend == "Strong" else (ORANGE if adx_trend == "Moderate" else GRAY)
        style_cell(ws, row, 7, adx_trend,
                   font=Font(color=adx_color, bold=True, size=10),
                   alignment=Alignment(horizontal="center"))

        # RSI
        style_cell(ws, row, 8, t.get("rsi", "N/A"),
                   num_fmt="0.0" if t.get("rsi") is not None else None,
                   alignment=Alignment(horizontal="center"))

        # RSI State
        rsi_state = t.get("rsi_state", "N/A")
        rsi_color = RED if rsi_state == "Overbought" else (GREEN if rsi_state == "Oversold" else GRAY)
        style_cell(ws, row, 9, rsi_state,
                   font=Font(color=rsi_color, size=10),
                   alignment=Alignment(horizontal="center"))

        # Trend Regime
        trend_regime = t.get("trend_regime", "N/A")
        style_cell(ws, row, 10, trend_regime,
                   font=_trend_font(trend_regime),
                   alignment=Alignment(horizontal="center"))
        row += 1

    row += 1

    # --- Sentiment indicators (label-value) ---
    # VIX
    vix_info = market_data.get("vix", {})
    vix_val = vix_info.get("current", 0)
    vix_level = vix_info.get("level", "N/A")
    if vix_val >= 30:
        vix_color = RED
    elif vix_val >= 20:
        vix_color = ORANGE
    elif vix_val <= 12:
        vix_color = GREEN
    else:
        vix_color = MED_BLUE

    row = write_label_value(ws, row, "VIX",
                            f"{vix_val:.2f}  ({vix_level})",
                            val_font=Font(bold=True, color=vix_color, size=11))

    # Fear & Greed
    fg = market_data.get("fear_greed", {})
    fg_score = fg.get("score", 0)
    fg_level = fg.get("level", "N/A")
    if fg_score >= 60:
        fg_color = GREEN
    elif fg_score <= 40:
        fg_color = RED
    else:
        fg_color = ORANGE
    row = write_label_value(ws, row, "Fear & Greed",
                            f"{fg_score:.1f}  ({fg_level})",
                            val_font=Font(bold=True, color=fg_color, size=11))

    # Put/Call Ratio (optional)
    putcall = market_data.get("putcall")
    if putcall:
        pc_ratio = putcall.get("ratio", "N/A")
        pc_sent = putcall.get("sentiment", "")
        row = write_label_value(ws, row, "Put/Call Ratio",
                                f"{pc_ratio}  ({pc_sent})" if pc_sent else str(pc_ratio))

    # CNN Fear & Greed (optional)
    cnn_fg = market_data.get("cnn_fg")
    if cnn_fg:
        cnn_score = cnn_fg.get("score", "N/A")
        cnn_level = cnn_fg.get("level", "")
        row = write_label_value(ws, row, "CNN Fear & Greed",
                                f"{cnn_score}  ({cnn_level})" if cnn_level else str(cnn_score))

    row += 1

    # --- Market Diagnosis ---
    diagnosis = market_data.get("diagnosis", {})
    verdict = diagnosis.get("verdict", "N/A")
    if "Bullish" in verdict:
        verdict_color = GREEN
    elif "Bearish" in verdict:
        verdict_color = RED
    else:
        verdict_color = ORANGE

    row = write_label_value(ws, row, "Verdict", verdict,
                            val_font=Font(bold=True, size=14, color=verdict_color))

    description = diagnosis.get("description", "")
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=10)
    desc_cell = ws.cell(row=row, column=1, value=description)
    desc_cell.font = Font(size=10)
    desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
    row += 3

    bullish_n = diagnosis.get("bullish_signals", 0)
    bearish_n = diagnosis.get("bearish_signals", 0)
    row = write_label_value(ws, row, "Bullish / Bearish Signals",
                            f"{bullish_n} / {bearish_n}")

    # --- Portfolio Exposure Dashboard ---
    exposure = market_data.get("exposure")
    if exposure:
        row += 1
        row = _build_exposure_dashboard(ws, row, exposure)

    return row


# ─── Portfolio Exposure Dashboard (inside Section 1) ──────────


def _component_color(raw_score: int) -> str:
    """Return a hex color for a component raw score.

    Args:
        raw_score: Integer raw score (negative / zero / positive).

    Returns:
        Hex color string.
    """
    if raw_score > 0:
        return GREEN
    if raw_score < 0:
        return RED
    return GRAY


def _component_fill(raw_score: int) -> PatternFill:
    """Return a pastel background fill for a component raw score.

    Args:
        raw_score: Integer raw score.

    Returns:
        PatternFill with pastel background.
    """
    if raw_score > 0:
        return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    if raw_score < 0:
        return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    return PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def _build_exposure_dashboard(
    ws: Any,
    start_row: int,
    exposure: dict[str, Any],
) -> int:
    """Build Portfolio Exposure sub-section within Section 1.

    Displays the composite regime score, 4 component breakdowns,
    and recommended Net/Gross/Long/Short allocation.

    Args:
        ws: Worksheet.
        start_row: Starting row.
        exposure: Dict from compute_regime_composite().

    Returns:
        Next available row.
    """
    row = start_row

    # Sub-section header
    row = write_section_header(ws, row, "Portfolio Exposure", cols=10)

    # Regime label + Composite Score
    composite = exposure.get("composite_score", 0)
    regime_label = exposure.get("regime_label", "N/A")
    regime_en = exposure.get("regime_label_en", "N/A")

    if composite >= 3:
        regime_color = GREEN
    elif composite >= 0:
        regime_color = ORANGE
    else:
        regime_color = RED

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    style_cell(ws, row, 1,
               f"Regime: {regime_label} ({regime_en})",
               font=Font(bold=True, size=12, color=regime_color))
    for c in range(2, 6):
        ws.cell(row=row, column=c).border = THIN_BORDER

    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=10)
    style_cell(ws, row, 6,
               f"Composite Score: {composite:+.1f}",
               font=Font(bold=True, size=12, color=regime_color),
               alignment=Alignment(horizontal="right"))
    for c in range(7, 11):
        ws.cell(row=row, column=c).border = THIN_BORDER
    row += 1

    # Component breakdown — 4 blocks, each 2 cols, with 1-col gap
    # Layout: cols 1-2 (Trend), 3-4 (Breadth), 5-6 (VIX), 7-8 (Sentiment)
    components = exposure.get("components", {})
    comp_order = [
        ("Trend", "trend"),
        ("Breadth", "breadth"),
        ("VIX", "vix"),
        ("Sentiment", "sentiment"),
    ]
    col_starts = [1, 3, 5, 7]

    # Row 1: Component name headers
    for (label, _key), start_col in zip(comp_order, col_starts):
        end_col = start_col + 1
        ws.merge_cells(
            start_row=row, start_column=start_col,
            end_row=row, end_column=end_col,
        )
        style_cell(ws, row, start_col, label,
                   font=Font(bold=True, color="FFFFFF", size=10),
                   fill=HDR_FILL,
                   alignment=Alignment(horizontal="center"))
        ws.cell(row=row, column=end_col).fill = HDR_FILL
        ws.cell(row=row, column=end_col).border = THIN_BORDER
    row += 1

    # Row 2: raw (×weight)
    for (_label, key), start_col in zip(comp_order, col_starts):
        comp = components.get(key, {})
        raw = comp.get("raw", 0)
        weight = comp.get("weight", 0)
        end_col = start_col + 1

        ws.merge_cells(
            start_row=row, start_column=start_col,
            end_row=row, end_column=end_col,
        )
        style_cell(ws, row, start_col,
                   f"{raw:+d} (×{weight:.1f})",
                   font=Font(bold=True, size=10, color=_component_color(raw)),
                   fill=_component_fill(raw),
                   alignment=Alignment(horizontal="center"))
        ws.cell(row=row, column=end_col).fill = _component_fill(raw)
        ws.cell(row=row, column=end_col).border = THIN_BORDER
    row += 1

    # Row 3: = weighted
    for (_label, key), start_col in zip(comp_order, col_starts):
        comp = components.get(key, {})
        raw = comp.get("raw", 0)
        weighted = comp.get("weighted", 0)
        end_col = start_col + 1

        ws.merge_cells(
            start_row=row, start_column=start_col,
            end_row=row, end_column=end_col,
        )
        style_cell(ws, row, start_col,
                   f"= {weighted:+.1f}",
                   font=Font(bold=True, size=11, color=_component_color(raw)),
                   fill=_component_fill(raw),
                   alignment=Alignment(horizontal="center"))
        ws.cell(row=row, column=end_col).fill = _component_fill(raw)
        ws.cell(row=row, column=end_col).border = THIN_BORDER
    row += 2

    # Net / Gross exposure
    net = exposure.get("net_exposure", 0)
    gross = exposure.get("gross_exposure", 0)
    net_color = GREEN if net > 0 else (RED if net < 0 else GRAY)
    row = write_label_value(
        ws, row, "Net Exposure",
        f"{net:.0f}%",
        val_font=Font(bold=True, size=12, color=net_color),
    )

    # Append Gross on same conceptual block
    row = write_label_value(
        ws, row, "Gross Exposure",
        f"{gross:.0f}%",
        val_font=Font(bold=True, size=11, color=MED_BLUE),
    )

    # Long / Short allocation
    long_alloc = exposure.get("long_allocation", 0)
    short_alloc = exposure.get("short_allocation", 0)
    row = write_label_value(
        ws, row, "Long Allocation",
        f"{long_alloc:.1f}%",
        val_font=Font(bold=True, size=11, color=GREEN),
    )
    row = write_label_value(
        ws, row, "Short Allocation",
        f"{short_alloc:.1f}%",
        val_font=Font(bold=True, size=11, color=RED),
    )

    # Breadth
    breadth = exposure.get("breadth_pct", 0)
    breadth_color = GREEN if breadth > 70 else (RED if breadth < 30 else ORANGE)
    row = write_label_value(
        ws, row, "Breadth (above 200 DMA)",
        f"{breadth:.1f}%",
        val_font=Font(bold=True, size=11, color=breadth_color),
    )

    return row


# ─── Section 2: Stock Trend & Technical ───────────────────────


def _build_stock_section(
    ws: Any,
    start_row: int,
    stock_data: list[dict[str, Any]],
) -> int:
    """Build Section 2: per-stock trend and technical indicators.

    Args:
        ws: Worksheet.
        start_row: Starting row.
        stock_data: List of stock dicts sorted by market cap.

    Returns:
        Next available row.
    """
    row = start_row

    # Section header
    row = write_section_header(ws, row, "Section 2: Stock Trend & Technical", cols=17)

    # Trend color legend row
    # Layout: 6 items spread across 16 columns — each gets ~2-3 merged cols
    legend_spans = [(1, 3), (4, 6), (7, 9), (10, 12), (13, 15), (16, 17)]
    for (start_col, end_col), (regime, kr_label) in zip(legend_spans, _TREND_LEGEND):
        if start_col != end_col:
            ws.merge_cells(
                start_row=row, start_column=start_col,
                end_row=row, end_column=end_col,
            )
        cell = ws.cell(row=row, column=start_col)
        cell.value = f"{regime} ({kr_label})"
        cell.font = Font(size=9, bold=True, color=_TREND_STYLES[regime][0])
        cell.fill = _trend_fill(regime)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 18
    row += 1

    # Column headers
    stk_headers = [
        ("No", 5), ("Ticker", 8), ("Name", 18), ("Sector", 14), ("Price", 12),
        ("1D%", 8), ("1W%", 8), ("Mkt Cap", 14), ("RSI", 7),
        ("RSI State", 11), ("MACD Sig", 10), ("ST Dir", 10),
        ("ADX", 7), ("ADX Trend", 11), ("Trend", 14), ("Tech Score", 11),
        ("Pattern", 14),
    ]
    for col_idx, (header, width) in enumerate(stk_headers, 1):
        style_cell(ws, row, col_idx, header, font=HDR_FONT, fill=HDR_FILL,
                   alignment=Alignment(horizontal="center"))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(
            ws.column_dimensions[col_letter].width or 0, width,
        )
    row += 1

    if not stock_data:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
        ws.cell(row=row, column=1, value="No data available").font = Font(
            size=10, color=GRAY, italic=True)
        return row + 1

    for i, s in enumerate(stock_data):
        regime = s.get("trend_regime", "N/A")
        ticker_fill = _trend_fill(regime)

        style_cell(ws, row, 1, i + 1, alignment=Alignment(horizontal="center"))
        style_cell(ws, row, 2, s["ticker"],
                   font=Font(bold=True, size=10, color=_TREND_STYLES.get(regime, (GRAY, False))[0]),
                   fill=ticker_fill)
        style_cell(ws, row, 3, s.get("name", ""))

        # Sector
        style_cell(ws, row, 4, s.get("sector", "N/A"),
                   font=Font(size=10, color=GRAY),
                   alignment=Alignment(horizontal="center"))

        style_cell(ws, row, 5, s.get("price", 0), num_fmt="#,##0.00",
                   alignment=Alignment(horizontal="right"))

        # 1D%
        chg_1d = s.get("change_1d")
        if chg_1d is not None:
            style_cell(ws, row, 6, chg_1d / 100, font=pct_color_font(chg_1d),
                       num_fmt="+0.00%;-0.00%",
                       alignment=Alignment(horizontal="center"))
        else:
            style_cell(ws, row, 6, "N/A", alignment=Alignment(horizontal="center"))

        # 1W%
        chg_1w = s.get("change_1w")
        if chg_1w is not None:
            style_cell(ws, row, 7, chg_1w / 100, font=pct_color_font(chg_1w),
                       num_fmt="+0.00%;-0.00%",
                       alignment=Alignment(horizontal="center"))
        else:
            style_cell(ws, row, 7, "N/A", alignment=Alignment(horizontal="center"))

        # Mkt Cap
        mc = s.get("market_cap")
        style_cell(ws, row, 8, format_market_cap(mc),
                   alignment=Alignment(horizontal="right"))

        # RSI + color
        rsi = s.get("rsi")
        if rsi is not None:
            rsi_font_color = RED if rsi >= 70 else (GREEN if rsi <= 30 else GRAY)
            style_cell(ws, row, 9, rsi, num_fmt="0.0",
                       font=Font(color=rsi_font_color, size=10),
                       alignment=Alignment(horizontal="center"))
        else:
            style_cell(ws, row, 9, "N/A", alignment=Alignment(horizontal="center"))

        # RSI State
        rsi_state = s.get("rsi_state", "N/A")
        rsi_st_color = RED if rsi_state == "Overbought" else (GREEN if rsi_state == "Oversold" else GRAY)
        style_cell(ws, row, 10, rsi_state,
                   font=Font(color=rsi_st_color, size=10),
                   alignment=Alignment(horizontal="center"))

        # MACD Signal
        macd_sig = s.get("macd_signal", "N/A")
        macd_color = GREEN if macd_sig == "Bullish" else (RED if macd_sig == "Bearish" else GRAY)
        style_cell(ws, row, 11, macd_sig,
                   font=Font(color=macd_color, bold=True, size=10),
                   alignment=Alignment(horizontal="center"))

        # Supertrend Direction
        st_dir = s.get("supertrend_dir", "N/A")
        st_color = GREEN if st_dir == "Bullish" else (RED if st_dir == "Bearish" else GRAY)
        style_cell(ws, row, 12, st_dir,
                   font=Font(color=st_color, bold=True, size=10),
                   alignment=Alignment(horizontal="center"))

        # ADX
        adx = s.get("adx")
        if adx is not None:
            style_cell(ws, row, 13, adx, num_fmt="0.0",
                       alignment=Alignment(horizontal="center"))
        else:
            style_cell(ws, row, 13, "N/A", alignment=Alignment(horizontal="center"))

        # ADX Trend
        adx_trend = s.get("adx_trend", "N/A")
        adx_color = GREEN if adx_trend == "Strong" else (ORANGE if adx_trend == "Moderate" else GRAY)
        style_cell(ws, row, 14, adx_trend,
                   font=Font(color=adx_color, bold=True, size=10),
                   alignment=Alignment(horizontal="center"))

        # Trend Regime
        trend_regime = s.get("trend_regime", "N/A")
        style_cell(ws, row, 15, trend_regime,
                   font=_trend_font(trend_regime),
                   alignment=Alignment(horizontal="center"))

        # Tech Score
        ts = s.get("tech_score", 0)
        style_cell(ws, row, 16, ts, num_fmt="0.0",
                   fill=score_fill(ts),
                   alignment=Alignment(horizontal="center"))

        # Pattern
        pattern = s.get("pattern", "—")
        pattern_style = _PATTERN_STYLES.get(pattern)
        if pattern_style:
            p_font_color, p_fill_color = pattern_style
            style_cell(ws, row, 17, pattern,
                       font=Font(bold=True, color=p_font_color, size=10),
                       fill=PatternFill(
                           start_color=p_fill_color,
                           end_color=p_fill_color,
                           fill_type="solid",
                       ),
                       alignment=Alignment(horizontal="center"))
        else:
            style_cell(ws, row, 17, pattern,
                       font=Font(color=GRAY, size=10),
                       alignment=Alignment(horizontal="center"))

        row += 1

    return row


# ─── Section 3: Sigma Bands ──────────────────────────────────


def _classify_band_position(
    price: float,
    s1_lower: float,
    s1_upper: float,
    s2_lower: float,
    s2_upper: float,
) -> str:
    """Classify where the current price sits within sigma bands.

    Args:
        price: Current stock price.
        s1_lower: 1-sigma lower bound.
        s1_upper: 1-sigma upper bound.
        s2_lower: 2-sigma lower bound.
        s2_upper: 2-sigma upper bound.

    Returns:
        Position label in Korean.
    """
    if price <= s2_lower:
        return "2σ하단 이탈"
    mid = (s1_lower + s1_upper) / 2
    range_1s = s1_upper - s1_lower
    if range_1s == 0:
        return "밴드 중간"

    # Near 2σ lower (between 2σ lower and 1σ lower)
    if price < s1_lower:
        return "2σ하단 근접"
    # Near 1σ lower (between 1σ lower and midpoint - threshold)
    if price < mid - range_1s * 0.15:
        return "1σ하단 근접"
    # Near 1σ upper (between midpoint + threshold and 1σ upper)
    if price > mid + range_1s * 0.15 and price <= s1_upper:
        return "1σ상단 근접"
    # Near 2σ upper (above 1σ upper)
    if price > s1_upper:
        return "2σ상단 근접"
    return "밴드 중간"


def _build_sigma_section(
    ws: Any,
    start_row: int,
    sigma_data: list[dict[str, Any]],
) -> int:
    """Build Section 3: sigma band analysis per stock.

    Each stock occupies 2 rows (1-week and 1-month periods).

    Args:
        ws: Worksheet.
        start_row: Starting row.
        sigma_data: List of analyze_ticker() result dicts.

    Returns:
        Next available row.
    """
    row = start_row

    if not sigma_data:
        return row

    # Section header
    row = write_section_header(ws, row, "Section 3: Sigma Bands (Option Volatility)", cols=12)

    # Column headers
    sig_headers = [
        ("Ticker", 8), ("Price", 11), ("HV20%", 8), ("Period", 8),
        ("Source", 8), ("Vol%", 8), ("1σ Lower", 11), ("1σ Upper", 11),
        ("2σ Lower", 11), ("2σ Upper", 11), ("±1σ%", 8), ("±2σ%", 8),
    ]
    for col_idx, (header, width) in enumerate(sig_headers, 1):
        style_cell(ws, row, col_idx, header, font=HDR_FONT, fill=HDR_FILL,
                   alignment=Alignment(horizontal="center"))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(
            ws.column_dimensions[col_letter].width or 0, width,
        )
    row += 1

    ok_results = [r for r in sigma_data if r.get("status") == "ok"]
    if not ok_results:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        ws.cell(row=row, column=1, value="No sigma data available").font = Font(
            size=10, color=GRAY, italic=True)
        return row + 1

    alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

    for idx, result in enumerate(ok_results):
        ticker = result["ticker"]
        price = result["current_price"]
        hv20 = result.get("hv20")
        periods = result.get("periods", {})
        use_alt = idx % 2 == 1  # alternate row shading per ticker

        # 티커 색상: 첫 번째 기간(주간)의 pos_1s 기준 5단계
        first_band = next(
            (b for b in periods.values() if "sigma_1" in b), {},
        )
        tk_pos = first_band.get("pos_1s", 0.0)
        if tk_pos > 100:
            tk_color = GREEN        # 진한초록: +1σ~+2σ
        elif tk_pos > 0:
            tk_color = "70AD47"     # 연한초록: 0~+1σ
        elif tk_pos == 0:
            tk_color = GRAY         # 회색: 중립
        elif tk_pos >= -100:
            tk_color = ORANGE       # 주황: 0~-1σ
        else:
            tk_color = RED          # 빨강: -1σ~-2σ

        first_period = True
        for period_label, band in periods.items():
            if "error" in band:
                continue

            bg_fill = alt_fill if use_alt else None

            # Ticker + Price only on first row for this ticker
            if first_period:
                style_cell(ws, row, 1, ticker,
                           font=Font(bold=True, size=10, color=tk_color),
                           fill=bg_fill)
                style_cell(ws, row, 2, price, num_fmt="$#,##0.00",
                           alignment=Alignment(horizontal="right"),
                           fill=bg_fill)
                style_cell(ws, row, 3,
                           f"{hv20:.1f}" if hv20 is not None else "N/A",
                           alignment=Alignment(horizontal="center"),
                           fill=bg_fill)
                first_period = False
            else:
                style_cell(ws, row, 1, "", fill=bg_fill)
                style_cell(ws, row, 2, "", fill=bg_fill)
                style_cell(ws, row, 3, "", fill=bg_fill)

            # Period
            style_cell(ws, row, 4, period_label,
                       alignment=Alignment(horizontal="center"),
                       fill=bg_fill)

            # Source (IV = blue, HV20 = gray)
            vol_source = band.get("vol_source", "?")
            src_font = _IV_FONT if vol_source == "IV" else _HV_FONT
            style_cell(ws, row, 5, vol_source, font=src_font,
                       alignment=Alignment(horizontal="center"),
                       fill=bg_fill)

            # Vol%
            style_cell(ws, row, 6, band.get("vol_pct", 0), num_fmt="0.0",
                       alignment=Alignment(horizontal="center"),
                       fill=bg_fill)

            # Sigma bounds
            s1 = band.get("sigma_1", {})
            s2 = band.get("sigma_2", {})
            s1_lower = s1.get("lower", 0)
            s1_upper = s1.get("upper", 0)
            s2_lower = s2.get("lower", 0)
            s2_upper = s2.get("upper", 0)

            style_cell(ws, row, 7, s1_lower, num_fmt="$#,##0.00",
                       alignment=Alignment(horizontal="right"),
                       fill=bg_fill)
            style_cell(ws, row, 8, s1_upper, num_fmt="$#,##0.00",
                       alignment=Alignment(horizontal="right"),
                       fill=bg_fill)
            style_cell(ws, row, 9, s2_lower, num_fmt="$#,##0.00",
                       alignment=Alignment(horizontal="right"),
                       fill=bg_fill)
            style_cell(ws, row, 10, s2_upper, num_fmt="$#,##0.00",
                       alignment=Alignment(horizontal="right"),
                       fill=bg_fill)

            # ±1σ% / ±2σ% 범위 표시 + 위치 기반 색상
            sigma_move_pct = band.get("sigma_move_pct", 0)
            sigma_2_pct = round(sigma_move_pct * 2, 2)
            pos_1s = band.get("pos_1s", 0.0)

            # 1σ 색상: 0~+1σ 구간 = 초록, 0~-1σ 구간 = 빨강
            color_1s = GREEN if pos_1s > 0 else RED if pos_1s < 0 else GRAY
            # 2σ 색상: +1σ~+2σ 구간 = 초록, -1σ~-2σ 구간 = 빨강, 1σ 안쪽이면 회색
            if pos_1s > 100:
                color_2s = GREEN
            elif pos_1s < -100:
                color_2s = RED
            else:
                color_2s = GRAY

            style_cell(ws, row, 11, f"±{sigma_move_pct:.1f}%",
                       font=Font(size=10, bold=True, color=color_1s),
                       alignment=Alignment(horizontal="center"),
                       fill=bg_fill)
            style_cell(ws, row, 12, f"±{sigma_2_pct:.1f}%",
                       font=Font(size=10, bold=True, color=color_2s),
                       alignment=Alignment(horizontal="center"),
                       fill=bg_fill)

            row += 1

    return row


# ─── Guide Sheet ────────────────────────────────────────────


# Guide content: list of (section_title, rows) where each row is
# (col_A, col_B, col_C) or a merged description string.
_GUIDE_SECTIONS: list[tuple[str, list[tuple[str, ...]]]] = [
    ("1. ADX (Average Directional Index) — 추세 강도", [
        ("ADX 값", "ADX Trend", "의미"),
        ("25 이상", "Strong", "뚜렷한 추세 존재 (상승이든 하락이든 강하게 움직이는 중)"),
        ("20 ~ 25", "Moderate", "약한 추세 형성 중 (방향이 잡히려는 단계)"),
        ("20 미만", "Weak", "추세 없음 — 횡보 구간"),
    ]),
    ("2. Supertrend (ST Dir) — 추세 방향", [
        ("값", "", "의미"),
        ("Bullish", "", "가격이 Supertrend 라인 위 — 상승 추세"),
        ("Bearish", "", "가격이 Supertrend 라인 아래 — 하락 추세"),
    ]),
    ("3. RSI (Relative Strength Index) — 과매수/과매도", [
        ("RSI 값", "상태", "의미"),
        ("70 이상", "Overbought", "과매수 — 단기 조정 가능성"),
        ("30 이하", "Oversold", "과매도 — 반등 가능성"),
        ("30 ~ 70", "Neutral", "중립 구간"),
    ]),
    ("4. Trend Regime — 종합 추세 분류 (ADX + EMA + DI)", [
        ("분류", "조건", "의미"),
        ("Strong Uptrend", "ADX ≥ 25, +DI > -DI, EMA 정배열",
         "강한 상승 추세. 매수 포지션 유리."),
        ("Uptrend", "ADX 20~25, +DI > -DI, EMA 정배열",
         "약한 상승 추세. 상승 초기 또는 약해지는 상승."),
        ("Sideways", "ADX < 20",
         "횡보. 추세가 없어서 방향 판단 불가. 돌파 대기."),
        ("Transition", "ADX ≥ 20이지만 DI↔EMA 방향 불일치",
         "전환 구간. 곧 방향이 결정됨. 상승/하락 양쪽 가능."),
        ("Downtrend", "ADX 20~25, +DI < -DI, EMA 역배열",
         "약한 하락 추세. 하락 초기 또는 약해지는 하락."),
        ("Strong Downtrend", "ADX ≥ 25, +DI < -DI, EMA 역배열",
         "강한 하락 추세. 역추세 매수 위험."),
    ]),
    ("5. EMA 배열 (EMA 20/50/200) — 장기 구조", [
        ("배열", "조건", "의미"),
        ("정배열", "EMA20 > EMA50 > EMA200",
         "단기 이평선이 장기 위 — 건강한 상승 구조"),
        ("역배열", "EMA20 < EMA50 < EMA200",
         "단기 이평선이 장기 아래 — 하락 구조"),
        ("혼재", "순서가 섞임",
         "추세 전환 과도기. 정배열/역배열로 정리될 때까지 관망."),
    ]),
    ("6. MACD Signal — 단기 모멘텀 방향", [
        ("값", "", "의미"),
        ("Bullish", "", "MACD선이 Signal선 위 — 단기 상승 모멘텀"),
        ("Bearish", "", "MACD선이 Signal선 아래 — 단기 하락 모멘텀"),
    ]),
    ("7. Tech Score — 종합 기술 점수 (0~100)", [
        ("점수", "색상", "의미"),
        ("80 이상", "초록", "강한 매수 신호"),
        ("60 ~ 80", "파랑", "양호"),
        ("40 ~ 60", "노랑", "중립"),
        ("40 미만", "빨강", "약세 신호"),
    ]),
    ("8. VIX — 시장 공포 지수", [
        ("VIX 값", "상태", "의미"),
        ("30 이상", "High (Fear)", "시장 공포 극대. 변동성 매우 높음."),
        ("20 ~ 30", "Elevated (Caution)", "경계 구간. 불확실성 증가."),
        ("12 ~ 20", "Normal", "정상 범위."),
        ("12 미만", "Low (Complacent)", "시장 안일. 역사적으로 급락 전조일 수 있음."),
    ]),
    ("9. Fear & Greed — 시장 심리 지수 (0~100)", [
        ("점수", "레벨", "의미"),
        ("80 ~ 100", "Extreme Greed", "극도의 탐욕. 과열 — 조정 가능성 높음."),
        ("60 ~ 80", "Greed", "탐욕. 상승 분위기이나 주의 필요."),
        ("40 ~ 60", "Neutral", "중립."),
        ("20 ~ 40", "Fear", "공포. 하락 분위기."),
        ("0 ~ 20", "Extreme Fear", "극도의 공포. 역사적으로 저점 매수 기회."),
    ]),
    ("10. Pattern Classification — 종목 매매 패턴 분류", [
        ("패턴", "조건 (5개 중 ≥3)", "의미"),
        ("Breakout (돌파)", "ADX 상승(+3), 거래량 폭발(1.5x), MACD Bullish, ST Bullish, EMA200 위",
         "저점 횡보 후 추세 전환 시작. 초기 진입 기회."),
        ("Pullback (눌림목)", "Trend ≥ Uptrend, EMA 정배열, 1W% 음수, RSI 35~55, ST Bullish",
         "강한 상승 추세 중 일시 조정. 추세 추종 매수 기회."),
        ("Consolidation (횡보 수렴)", "ADX < 20, BB 스퀴즈(하위 25%), 거래량 감소, EMA200 지지, MACD ≈ 0",
         "오랜 횡보 후 변동성 축소. 돌파 대기 종목."),
        ("—", "조건 미충족",
         "위 3가지 패턴에 해당하지 않음. 관망."),
    ]),
    ("11. Sigma Bands — 옵션 변동성 기반 가격 범위", [
        ("용어", "", "의미"),
        ("IV (Implied Vol)", "",
         "옵션 시장이 예상하는 향후 변동성. 장중에만 제공."),
        ("HV20 (Historical Vol)", "",
         "최근 20일 실현 변동성. IV 미제공 시 대체 사용."),
        ("1σ 범위", "",
         "68% 확률로 가격이 머무를 범위."),
        ("2σ 범위", "",
         "95% 확률로 가격이 머무를 범위."),
        ("2σ하단 근접/이탈", "",
         "통계적 저점. 반등 가능성이 높은 구간."),
        ("2σ상단 근접", "",
         "통계적 고점. 단기 조정 가능성."),
    ]),
]

_GUIDE_TIP_TEXT = (
    "[ 활용 팁 ]\n\n"
    "1) Trend가 Strong Uptrend/Uptrend이면서 Tech Score 60+인 종목 → 상승 추세 확인 종목\n"
    "2) Transition 종목 → 방향 결정 임박. 다음 리포트에서 Trend가 어떻게 바뀌는지 주시\n"
    "3) Strong Downtrend 종목 → 역추세 매수 위험. EMA 배열이 '혼재'로 바뀔 때까지 대기\n"
    "4) Sideways + 정배열 종목 → ADX가 20 이상으로 올라가면 상승 추세 전환 가능성\n"
    "5) 시그마 밴드에서 2σ 하단 근접 + Fear & Greed ≤ 30 → 통계적 + 심리적 저점 겹침\n"
    "6) VIX 30+ → 시장 전체 리스크 높음. 개별 종목 분석보다 현금 비중 우선 점검\n"
    "7) Breakout 종목 → 거래량 폭발과 ADX 상승 동반 시 초기 추세 진입 기회\n"
    "8) Pullback 종목 → 상승 추세 유지 중 눌림. RSI 35~55 구간에서 매수 타이밍 탐색\n"
    "9) Consolidation 종목 → BB 스퀴즈 후 돌파 대기. Breakout 전환 모니터링 필요"
)


def _build_guide_sheet(wb: Workbook) -> None:
    """Build the Guide sheet with indicator explanations.

    Args:
        wb: Target Workbook.
    """
    ws = wb.create_sheet(title="Guide")

    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")

    # Title
    ws.merge_cells("A1:C1")
    title = ws.cell(row=1, column=1, value="Signal Report 지표 가이드")
    title.font = TITLE_FONT
    title.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 60

    row = 3

    for section_title, rows in _GUIDE_SECTIONS:
        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        style_cell(ws, row, 1, section_title,
                   font=SECTION_FONT, fill=SECTION_FILL)
        for c in range(2, 4):
            cell = ws.cell(row=row, column=c)
            cell.fill = SECTION_FILL
            cell.border = THIN_BORDER
        row += 1

        for i, cols in enumerate(rows):
            is_header = i == 0
            font = HDR_FONT if is_header else VALUE_FONT
            fill = HDR_FILL if is_header else None

            for col_idx, val in enumerate(cols, 1):
                al = center if is_header else wrap
                style_cell(ws, row, col_idx, val, font=font, fill=fill,
                           alignment=al)
            row += 1

        row += 1  # gap between sections

    # Tips section
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    style_cell(ws, row, 1, "활용 팁",
               font=SECTION_FONT, fill=SECTION_FILL)
    for c in range(2, 4):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECTION_FILL
        cell.border = THIN_BORDER
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row + 8, end_column=3)
    tip_cell = ws.cell(row=row, column=1, value=_GUIDE_TIP_TEXT)
    tip_cell.font = Font(size=11)
    tip_cell.alignment = Alignment(wrap_text=True, vertical="top")
    tip_cell.border = THIN_BORDER
    for r in range(row, row + 9):
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER

    ws.sheet_properties.tabColor = DARK_BLUE
