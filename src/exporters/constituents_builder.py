"""Top_Stocks sheet builder — US and KR watchlist constituents."""

from __future__ import annotations

from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from src.analyzers.technical import TechnicalAnalyzer
from src.exporters.base import (
    DARK_BLUE,
    GRAY,
    HDR_FILL,
    HDR_FONT,
    LINK_FONT,
    SECTION_FILL,
    SECTION_FONT,
    THIN_BORDER,
    TITLE_FONT,
    format_market_cap,
    pct_color_font,
    score_fill,
    style_cell,
)


def build_constituents_sheet(
    wb: Workbook,
    us_stocks: list[dict[str, Any]],
    kr_stocks: list[dict[str, Any]],
) -> None:
    """Build the Top_Stocks sheet with US and KR watchlist stock data.

    Args:
        wb: Target Workbook.
        us_stocks: List of dicts with keys:
            ticker, name, ohlcv, current, change_1d, market_cap, tech_score
        kr_stocks: Same structure for Korean stocks.
    """
    ws = wb.create_sheet(title="Top_Stocks")

    # Title
    ws.merge_cells("A1:I1")
    ws.cell(row=1, column=1, value="Watchlist Stocks — Top Constituents").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    # Back link
    back_cell = ws.cell(row=1, column=10, value="<< Overview")
    back_cell.font = LINK_FONT
    back_cell.hyperlink = "#Overview!A1"

    row = 3

    # --- US Section ---
    row = _build_stock_table(ws, row, "US Watchlist", us_stocks)
    row += 2

    # --- KR Section ---
    row = _build_stock_table(ws, row, "KR Watchlist", kr_stocks)

    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = "70AD47"


def _build_stock_table(
    ws: Any,
    start_row: int,
    section_title: str,
    stocks: list[dict[str, Any]],
) -> int:
    """Build a stock table section.

    Args:
        ws: Worksheet.
        start_row: Starting row.
        section_title: Section header text.
        stocks: List of stock data dicts.

    Returns:
        Next available row.
    """
    # Section header
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=9)
    style_cell(ws, start_row, 1, section_title, font=SECTION_FONT, fill=SECTION_FILL)
    for c in range(2, 10):
        ws.cell(row=start_row, column=c).fill = SECTION_FILL
        ws.cell(row=start_row, column=c).border = THIN_BORDER
    start_row += 1

    # Column headers
    headers = [
        ("No", 5), ("Ticker", 10), ("Name", 20), ("Price", 14),
        ("1D%", 8), ("1W%", 8), ("Market Cap", 16),
        ("Tech Score", 12), ("Link", 8),
    ]
    for col_idx, (header, width) in enumerate(headers, 1):
        style_cell(ws, start_row, col_idx, header, font=HDR_FONT, fill=HDR_FILL,
                   alignment=Alignment(horizontal="center"))
        ws.column_dimensions[chr(64 + col_idx)].width = max(
            ws.column_dimensions[chr(64 + col_idx)].width or 0, width
        )
    start_row += 1

    # Data rows
    if not stocks:
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=9)
        ws.cell(row=start_row, column=1, value="No data available").font = Font(
            size=10, color=GRAY, italic=True)
        return start_row + 1

    for i, stock in enumerate(stocks):
        row = start_row + i
        style_cell(ws, row, 1, i + 1, alignment=Alignment(horizontal="center"))
        style_cell(ws, row, 2, stock["ticker"], font=Font(bold=True, size=10))
        style_cell(ws, row, 3, stock["name"])
        style_cell(ws, row, 4, stock.get("current", 0), num_fmt="#,##0.00")

        # 1D change
        chg_1d = stock.get("change_1d")
        if chg_1d is not None:
            style_cell(ws, row, 5, chg_1d / 100,
                       font=pct_color_font(chg_1d),
                       num_fmt="0.00%",
                       alignment=Alignment(horizontal="center"))
        else:
            style_cell(ws, row, 5, "N/A", alignment=Alignment(horizontal="center"))

        # 1W change
        chg_1w = stock.get("change_1w")
        if chg_1w is not None:
            style_cell(ws, row, 6, chg_1w / 100,
                       font=pct_color_font(chg_1w),
                       num_fmt="0.00%",
                       alignment=Alignment(horizontal="center"))
        else:
            style_cell(ws, row, 6, "N/A", alignment=Alignment(horizontal="center"))

        # Market cap
        mc = stock.get("market_cap")
        style_cell(ws, row, 7, format_market_cap(mc),
                   alignment=Alignment(horizontal="right"))

        # Tech score
        ts = stock.get("tech_score", 0)
        style_cell(ws, row, 8, ts, num_fmt="0.0",
                   fill=score_fill(ts),
                   alignment=Alignment(horizontal="center"))

        # Link to existing analysis file (if present)
        link_path = stock.get("analysis_link")
        if link_path:
            link_cell = style_cell(ws, row, 9, "Open", font=LINK_FONT,
                                   alignment=Alignment(horizontal="center"))
        else:
            style_cell(ws, row, 9, "-", alignment=Alignment(horizontal="center"))

    return start_row + len(stocks)
