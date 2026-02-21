"""Individual index detail sheet builder with report + 5 charts."""

from __future__ import annotations

from datetime import datetime
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font

from src.analyzers.technical import _bbands, _macd, _rsi, _sma
from src.analyzers.trend import _adx, _supertrend
from src.exporters.base import (
    DARK_BLUE,
    GRAY,
    GREEN,
    HDR_FILL,
    HDR_FONT,
    LABEL_FONT,
    LINE_WIDTH_HEAVY,
    LINE_WIDTH_LIGHT,
    LINE_WIDTH_MEDIUM,
    LINK_FONT,
    ORANGE,
    RED,
    THIN_BORDER,
    TITLE_FONT,
    VALUE_FONT,
    add_reference_line_series,
    apply_axis_labels,
    apply_chart_gridlines,
    apply_x_axis_tick_interval,
    apply_y_axis_number_format,
    apply_y_axis_padding,
    style_cell,
    style_line_series,
    write_constant_column,
    write_label_value,
    write_section_header,
)


def build_index_detail_sheet(
    wb: Workbook,
    sheet_name: str,
    display_name: str,
    ohlcv: pd.DataFrame,
) -> None:
    """Build a detailed analysis sheet for a single market index.

    Includes:
        - Left panel (A:B): Report with current value, indicators, trend summary
        - Right panel (D+): Chart data + 5 charts
            1. Price + SMA(20,60) + Bollinger Bands
            2. RSI
            3. MACD + Histogram
            4. ADX + DI
            5. Volume

    Args:
        wb: Target Workbook.
        sheet_name: Sheet tab name.
        display_name: Full display name for the index.
        ohlcv: OHLCV DataFrame.
    """
    ws = wb.create_sheet(title=sheet_name)
    close = ohlcv["Close"]
    current = float(close.iloc[-1])

    has_volume = "Volume" in ohlcv.columns
    has_hlc = all(c in ohlcv.columns for c in ("High", "Low", "Close"))

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 3
    for col_letter in "DEFGHIJKLMNOPQRST":
        ws.column_dimensions[col_letter].width = 13

    # ===== REPORT SECTION (A:B) =====
    row = 1

    # Title
    ws.merge_cells(f"A{row}:B{row}")
    ws.cell(row=row, column=1, value=display_name).font = TITLE_FONT
    ws.cell(row=row, column=1).alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 30
    row += 1

    # Date range
    ws.cell(row=row, column=1,
            value=f"Analysis: {datetime.now().strftime('%Y-%m-%d')}").font = Font(size=9, color=GRAY)
    ws.cell(row=row, column=2,
            value=f"Data: {ohlcv.index[0].date()} ~ {ohlcv.index[-1].date()}").font = Font(size=9, color=GRAY)
    row += 2

    # Overview
    row = write_section_header(ws, row, "Overview")
    row = write_label_value(ws, row, "Current Value", f"{current:,.2f}",
                            val_font=Font(bold=True, size=12))

    if has_volume:
        volume = int(ohlcv["Volume"].iloc[-1])
        row = write_label_value(ws, row, "Volume", f"{volume:,}")
    row = write_label_value(ws, row, "Data Points", f"{len(ohlcv)} days")

    # Returns
    if len(ohlcv) >= 2:
        prev = float(close.iloc[-2])
        chg_1d = (current / prev - 1) * 100 if prev else 0
        row = write_label_value(ws, row, "1D Change", f"{chg_1d:+.2f}%")
    if len(ohlcv) >= 6:
        past = float(close.iloc[-6])
        chg_1w = (current / past - 1) * 100 if past else 0
        row = write_label_value(ws, row, "1W Change", f"{chg_1w:+.2f}%")
    if len(ohlcv) >= 22:
        past = float(close.iloc[-22])
        chg_1m = (current / past - 1) * 100 if past else 0
        row = write_label_value(ws, row, "1M Change", f"{chg_1m:+.2f}%")
    row += 1

    # Technical indicators
    row = write_section_header(ws, row, "Technical Indicators")
    rsi_series = _rsi(close, 14)
    rsi_val = _safe_last(rsi_series)
    row = write_label_value(ws, row, "RSI (14)",
                            f"{rsi_val:.1f}" if rsi_val is not None else "N/A")

    macd_result = _macd(close, 12, 26, 9)
    macd_val = _safe_last(macd_result["macd"])
    macd_sig = _safe_last(macd_result["signal"])
    macd_hist = _safe_last(macd_result["histogram"])
    row = write_label_value(ws, row, "MACD",
                            f"{macd_val:.4f}" if macd_val is not None else "N/A")
    row = write_label_value(ws, row, "MACD Signal",
                            f"{macd_sig:.4f}" if macd_sig is not None else "N/A")
    row = write_label_value(ws, row, "MACD Histogram",
                            f"{macd_hist:.4f}" if macd_hist is not None else "N/A")

    sma20_val = _safe_last(_sma(close, 20))
    sma60_val = _safe_last(_sma(close, 60))
    row = write_label_value(ws, row, "SMA 20",
                            f"{sma20_val:,.2f}" if sma20_val is not None else "N/A")
    row = write_label_value(ws, row, "SMA 60",
                            f"{sma60_val:,.2f}" if sma60_val is not None else "N/A")
    row += 1

    # Trend indicators
    if has_hlc:
        row = write_section_header(ws, row, "Trend Indicators")
        adx_df = _adx(ohlcv)
        adx_val = _safe_last(adx_df["ADX"])
        plus_di_val = _safe_last(adx_df["plus_di"])
        minus_di_val = _safe_last(adx_df["minus_di"])
        row = write_label_value(ws, row, "ADX",
                                f"{adx_val:.1f}" if adx_val is not None else "N/A")
        row = write_label_value(ws, row, "+DI",
                                f"{plus_di_val:.1f}" if plus_di_val is not None else "N/A")
        row = write_label_value(ws, row, "-DI",
                                f"{minus_di_val:.1f}" if minus_di_val is not None else "N/A")

        if adx_val is not None:
            if adx_val >= 25:
                trend_label = "Strong Trend"
            elif adx_val >= 20:
                trend_label = "Moderate Trend"
            else:
                trend_label = "Weak / No Trend"
            row = write_label_value(ws, row, "ADX Interpretation", trend_label)

        st_df = _supertrend(ohlcv)
        st_dir = st_df["direction"].dropna()
        if not st_dir.empty:
            direction = "Bullish" if int(st_dir.iloc[-1]) == 1 else "Bearish"
            row = write_label_value(ws, row, "Supertrend", direction)
        row += 1

    # Back link
    back_cell = ws.cell(row=1, column=14, value="<< Overview")
    back_cell.font = LINK_FONT
    back_cell.hyperlink = "#Overview!A1"

    # ===== CHART DATA (D+ columns) =====
    chart_data_row = 2
    chart_headers = [
        "Date", "Close", "SMA_20", "SMA_60", "BB_Upper", "BB_Lower",
        "Volume", "RSI_14", "MACD", "MACD_Signal", "MACD_Hist",
    ]
    # Add ADX headers if available
    if has_hlc:
        chart_headers.extend(["ADX", "+DI", "-DI"])

    for j, h in enumerate(chart_headers):
        col = 4 + j
        style_cell(ws, chart_data_row, col, h, font=HDR_FONT, fill=HDR_FILL,
                   alignment=Alignment(horizontal="center"))

    # Compute full series
    sma20 = _sma(close, 20)
    sma60 = _sma(close, 60)
    bb = _bbands(close, 20, 2)
    rsi_full = _rsi(close, 14)
    macd_full = _macd(close, 12, 26, 9)

    adx_full = None
    if has_hlc:
        adx_full = _adx(ohlcv)

    n = len(ohlcv)
    for i in range(n):
        r_idx = chart_data_row + 1 + i
        dt = ohlcv.index[i]
        date_val = dt.tz_localize(None) if hasattr(dt, 'tz_localize') and dt.tzinfo else dt

        ws.cell(row=r_idx, column=4, value=date_val).number_format = "YYYY-MM-DD"
        ws.cell(row=r_idx, column=5, value=_round_val(close.iloc[i]))

        ws.cell(row=r_idx, column=6, value=_round_val(sma20.iloc[i]))
        ws.cell(row=r_idx, column=7, value=_round_val(sma60.iloc[i]))
        ws.cell(row=r_idx, column=8, value=_round_val(bb["upper"].iloc[i]))
        ws.cell(row=r_idx, column=9, value=_round_val(bb["lower"].iloc[i]))

        vol = int(ohlcv["Volume"].iloc[i]) if has_volume else 0
        ws.cell(row=r_idx, column=10, value=vol)
        ws.cell(row=r_idx, column=11, value=_round_val(rsi_full.iloc[i]))
        ws.cell(row=r_idx, column=12, value=_round_val(macd_full["macd"].iloc[i], 4))
        ws.cell(row=r_idx, column=13, value=_round_val(macd_full["signal"].iloc[i], 4))
        ws.cell(row=r_idx, column=14, value=_round_val(macd_full["histogram"].iloc[i], 4))

        if adx_full is not None:
            ws.cell(row=r_idx, column=15, value=_round_val(adx_full["ADX"].iloc[i]))
            ws.cell(row=r_idx, column=16, value=_round_val(adx_full["plus_di"].iloc[i]))
            ws.cell(row=r_idx, column=17, value=_round_val(adx_full["minus_di"].iloc[i]))

    data_end_row = chart_data_row + n

    dates_ref = Reference(ws, min_col=4, min_row=chart_data_row + 1,
                          max_row=data_end_row)

    # --- Reference-line constant columns (R=18, S=19, T=20) ---
    write_constant_column(ws, 18, chart_data_row, n, 30, "RSI_30")
    write_constant_column(ws, 19, chart_data_row, n, 70, "RSI_70")
    if has_hlc:
        write_constant_column(ws, 20, chart_data_row, n, 25, "ADX_25")

    # --- Chart 1: Price + SMA + Bollinger Bands ---
    price_chart = LineChart()
    price_chart.title = f"{display_name} — Price & Moving Averages"
    price_chart.style = 10
    price_chart.width = 32
    price_chart.height = 16
    price_chart.legend.position = "b"
    apply_chart_gridlines(price_chart)
    apply_y_axis_number_format(price_chart, "#,##0.00")
    apply_axis_labels(price_chart, x_title="Date", y_title="Price",
                      x_numfmt="yyyy-mm")
    # Y-axis padding based on Close + BB range
    price_vals = pd.concat([close, bb["upper"], bb["lower"]]).dropna()
    if not price_vals.empty:
        apply_y_axis_padding(price_chart, float(price_vals.min()),
                             float(price_vals.max()))
    apply_x_axis_tick_interval(price_chart, n)

    series_defs = [
        (5, "Close", "2F5496", LINE_WIDTH_HEAVY, None),
        (6, "SMA_20", "ED7D31", LINE_WIDTH_MEDIUM, "dash"),
        (7, "SMA_60", "70AD47", LINE_WIDTH_MEDIUM, "dash"),
        (8, "BB_Upper", "A5A5A5", LINE_WIDTH_LIGHT, "dot"),
        (9, "BB_Lower", "A5A5A5", LINE_WIDTH_LIGHT, "dot"),
    ]
    for col, _, color, width, dash in series_defs:
        ref = Reference(ws, min_col=col, min_row=chart_data_row, max_row=data_end_row)
        price_chart.add_data(ref, titles_from_data=True)

    price_chart.set_categories(dates_ref)
    for idx, (_, _, color, width, dash) in enumerate(series_defs):
        style_line_series(price_chart.series[idx], color, width, dash)

    ws.add_chart(price_chart, "D1")

    # --- Chart 2: RSI with 30/70 reference lines ---
    rsi_chart = LineChart()
    rsi_chart.title = "RSI (14)"
    rsi_chart.style = 10
    rsi_chart.width = 32
    rsi_chart.height = 12
    rsi_chart.y_axis.scaling.min = 0
    rsi_chart.y_axis.scaling.max = 100
    rsi_chart.legend.position = "b"
    apply_chart_gridlines(rsi_chart)
    apply_axis_labels(rsi_chart, x_title="Date", y_title="RSI",
                      y_numfmt="0", x_numfmt="yyyy-mm")
    rsi_chart.y_axis.majorUnit = 10  # 0-100 / 10
    rsi_chart.y_axis.tickLblPos = "nextTo"
    rsi_chart.y_axis.delete = False
    apply_x_axis_tick_interval(rsi_chart, n)

    rsi_ref = Reference(ws, min_col=11, min_row=chart_data_row, max_row=data_end_row)
    rsi_chart.add_data(rsi_ref, titles_from_data=True)
    rsi_chart.set_categories(dates_ref)
    style_line_series(rsi_chart.series[0], "7030A0", LINE_WIDTH_HEAVY)

    # RSI 30 / 70 reference lines
    add_reference_line_series(rsi_chart, ws, 18, chart_data_row, data_end_row, RED)
    add_reference_line_series(rsi_chart, ws, 19, chart_data_row, data_end_row, RED)

    ws.add_chart(rsi_chart, "D19")

    # --- Chart 3: MACD + Histogram (green/red) ---
    macd_chart = LineChart()
    macd_chart.title = "MACD (12, 26, 9)"
    macd_chart.style = 10
    macd_chart.width = 32
    macd_chart.height = 12
    macd_chart.legend.position = "b"
    apply_chart_gridlines(macd_chart)
    apply_axis_labels(macd_chart, x_title="Date", y_title="MACD",
                      y_numfmt="0.0000", x_numfmt="yyyy-mm")
    macd_chart.y_axis.tickLblPos = "nextTo"
    macd_chart.y_axis.delete = False
    apply_x_axis_tick_interval(macd_chart, n)

    macd_line_ref = Reference(ws, min_col=12, min_row=chart_data_row, max_row=data_end_row)
    macd_chart.add_data(macd_line_ref, titles_from_data=True)
    macd_sig_ref = Reference(ws, min_col=13, min_row=chart_data_row, max_row=data_end_row)
    macd_chart.add_data(macd_sig_ref, titles_from_data=True)
    macd_chart.set_categories(dates_ref)
    style_line_series(macd_chart.series[0], "2F5496", LINE_WIDTH_HEAVY)
    style_line_series(macd_chart.series[1], "ED7D31", LINE_WIDTH_MEDIUM, "dash")

    hist_bar = BarChart()
    hist_ref = Reference(ws, min_col=14, min_row=chart_data_row, max_row=data_end_row)
    hist_bar.add_data(hist_ref, titles_from_data=True)
    hist_bar.y_axis.axId = 200

    # Color each histogram bar green (positive) or red (negative)
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    for i in range(n):
        r_idx = chart_data_row + 1 + i
        cell_val = ws.cell(row=r_idx, column=14).value
        pt = DataPoint(idx=i)
        if cell_val is not None and cell_val >= 0:
            pt.graphicalProperties.solidFill = "548235"  # green
        else:
            pt.graphicalProperties.solidFill = "C00000"  # red
        hist_bar.series[0].data_points.append(pt)

    macd_chart += hist_bar

    ws.add_chart(macd_chart, "D33")

    # --- Chart 4: ADX + DI with 25 reference line ---
    if has_hlc:
        adx_chart = LineChart()
        adx_chart.title = "ADX & Directional Indicators"
        adx_chart.style = 10
        adx_chart.width = 32
        adx_chart.height = 12
        adx_chart.legend.position = "b"
        apply_chart_gridlines(adx_chart)
        apply_axis_labels(adx_chart, x_title="Date", y_title="ADX / DI",
                          y_numfmt="0.0", x_numfmt="yyyy-mm")
        adx_chart.y_axis.tickLblPos = "nextTo"
        adx_chart.y_axis.delete = False
        apply_x_axis_tick_interval(adx_chart, n)

        adx_ref = Reference(ws, min_col=15, min_row=chart_data_row, max_row=data_end_row)
        adx_chart.add_data(adx_ref, titles_from_data=True)
        plus_di_ref = Reference(ws, min_col=16, min_row=chart_data_row, max_row=data_end_row)
        adx_chart.add_data(plus_di_ref, titles_from_data=True)
        minus_di_ref = Reference(ws, min_col=17, min_row=chart_data_row, max_row=data_end_row)
        adx_chart.add_data(minus_di_ref, titles_from_data=True)
        adx_chart.set_categories(dates_ref)

        style_line_series(adx_chart.series[0], "2F5496", LINE_WIDTH_HEAVY)
        style_line_series(adx_chart.series[1], "70AD47", LINE_WIDTH_LIGHT, "dash")
        style_line_series(adx_chart.series[2], "C00000", LINE_WIDTH_LIGHT, "dash")

        # ADX 25 reference line
        add_reference_line_series(adx_chart, ws, 20, chart_data_row, data_end_row, ORANGE)

        ws.add_chart(adx_chart, "D47")

    # --- Chart 5: Volume ---
    if has_volume:
        vol_chart_anchor = "D61" if has_hlc else "D47"
        vol_chart = BarChart()
        vol_chart.title = "Volume"
        vol_chart.style = 10
        vol_chart.width = 32
        vol_chart.height = 10
        vol_chart.legend = None
        apply_chart_gridlines(vol_chart)
        apply_y_axis_number_format(vol_chart, "#,##0")
        apply_axis_labels(vol_chart, x_title="Date", y_title="Volume",
                          x_numfmt="yyyy-mm")
        vol_chart.y_axis.tickLblPos = "nextTo"
        vol_chart.y_axis.delete = False
        apply_x_axis_tick_interval(vol_chart, n)

        vol_ref = Reference(ws, min_col=10, min_row=chart_data_row, max_row=data_end_row)
        vol_chart.add_data(vol_ref, titles_from_data=True)
        vol_chart.set_categories(dates_ref)
        vol_chart.series[0].graphicalProperties.solidFill = "4472C4"

        ws.add_chart(vol_chart, vol_chart_anchor)

    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = "4472C4"


def _safe_last(series: pd.Series) -> float | None:
    """Get the last non-NaN value from a Series."""
    valid = series.dropna()
    if valid.empty:
        return None
    return float(valid.iloc[-1])


def _round_val(val: Any, decimals: int = 2) -> float | None:
    """Round a value, returning None for NaN."""
    if pd.isna(val):
        return None
    return round(float(val), decimals)
