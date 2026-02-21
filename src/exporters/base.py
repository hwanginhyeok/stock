"""Common Excel styles, colors, and utility functions for exporters."""

from __future__ import annotations

from openpyxl.chart.axis import ChartLines
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.worksheet import Worksheet

# ============================================================
# Color palette
# ============================================================

DARK_BLUE = "2F5496"
MED_BLUE = "4472C4"
LIGHT_BLUE = "D6E4F0"
GREEN = "548235"
LIGHT_GREEN = "C6EFCE"
RED = "C00000"
ORANGE = "ED7D31"
GRAY = "808080"
LIGHT_GRAY = "F2F2F2"
PURPLE = "7030A0"

# ============================================================
# Chart line-width constants (EMU: 1 pt = 12700 EMU)
# ============================================================

LINE_WIDTH_HEAVY = 28575       # 2.25pt — main data (Close, RSI, ADX, VIX)
LINE_WIDTH_MEDIUM = 22225      # 1.75pt — secondary indicators (SMA, MACD Signal)
LINE_WIDTH_LIGHT = 15875       # 1.25pt — background lines (BB, DI)
LINE_WIDTH_REFERENCE = 9525    # 0.75pt — reference/threshold lines (dashed)
REF_LINE_COLOR = "B0B0B0"

# ============================================================
# Reusable font/fill/border definitions
# ============================================================

HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
SECTION_FILL = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
LABEL_FONT = Font(bold=True, size=10)
VALUE_FONT = Font(size=10)
TITLE_FONT = Font(bold=True, size=14, color=DARK_BLUE)
SUBTITLE_FONT = Font(bold=True, size=12, color=DARK_BLUE)
LINK_FONT = Font(color="0563C1", underline="single", size=10)
SMALL_FONT = Font(size=9, color=GRAY)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
LABEL_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")


# ============================================================
# Utility functions
# ============================================================


def style_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: object,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    num_fmt: str | None = None,
    border: Border | None = None,
) -> object:
    """Write a value to a cell and apply formatting.

    Args:
        ws: Target worksheet.
        row: Row number (1-based).
        col: Column number (1-based).
        value: Cell value.
        font: Optional font.
        fill: Optional fill.
        alignment: Optional alignment.
        num_fmt: Optional number format string.
        border: Optional border (defaults to THIN_BORDER).

    Returns:
        The openpyxl Cell object.
    """
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if num_fmt:
        cell.number_format = num_fmt
    cell.border = border or THIN_BORDER
    return cell


def score_fill(score: float) -> PatternFill:
    """Return a PatternFill color based on a 0-100 score.

    Args:
        score: Numeric score.

    Returns:
        Green for >=80, blue for >=60, yellow for >=40, red otherwise.
    """
    if score >= 80:
        return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    if score >= 60:
        return PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    if score >= 40:
        return PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def pct_color_font(value: float, bold: bool = False) -> Font:
    """Return a colored Font for a percentage change value.

    Args:
        value: Percentage change.
        bold: Whether to make the font bold.

    Returns:
        Green font for positive, red for negative, gray for zero.
    """
    if value > 0:
        color = GREEN
    elif value < 0:
        color = RED
    else:
        color = GRAY
    return Font(color=color, bold=bold, size=10)


def format_market_cap(val: float | None) -> str:
    """Format market cap to human-readable string.

    Args:
        val: Market cap in raw number.

    Returns:
        Formatted string like "$3.45T", "890B", "1,234조".
    """
    if val is None or val == 0:
        return "N/A"
    if val >= 1e12:
        return f"${val / 1e12:.2f}T"
    if val >= 1e9:
        return f"${val / 1e9:.1f}B"
    if val >= 1e6:
        return f"${val / 1e6:.0f}M"
    return f"${val:,.0f}"


def write_section_header(ws: Worksheet, row: int, text: str, cols: int = 2) -> int:
    """Write a colored section header spanning multiple columns.

    Args:
        ws: Target worksheet.
        row: Row number.
        text: Header text.
        cols: Number of columns to merge.

    Returns:
        Next row number.
    """
    if cols > 1:
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=cols,
        )
    style_cell(ws, row, 1, text, font=SECTION_FONT, fill=SECTION_FILL)
    for c in range(2, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECTION_FILL
        cell.border = THIN_BORDER
    return row + 1


def write_label_value(
    ws: Worksheet,
    row: int,
    label: str,
    value: object,
    num_fmt: str | None = None,
    val_font: Font | None = None,
) -> int:
    """Write a label-value pair in columns A and B.

    Args:
        ws: Target worksheet.
        row: Row number.
        label: Label text in column A.
        value: Value in column B.
        num_fmt: Optional number format.
        val_font: Optional font for value cell.

    Returns:
        Next row number.
    """
    style_cell(ws, row, 1, label, font=LABEL_FONT, fill=LABEL_FILL)
    style_cell(ws, row, 2, value, font=val_font or VALUE_FONT, num_fmt=num_fmt)
    return row + 1


# ============================================================
# Chart helper functions
# ============================================================


def apply_chart_gridlines(chart: object) -> None:
    """Enable major gridlines on both Y and X axes.

    Args:
        chart: An openpyxl chart object.
    """
    chart.y_axis.majorGridlines = ChartLines()
    chart.x_axis.majorGridlines = ChartLines()


def apply_y_axis_number_format(chart: object, fmt: str) -> None:
    """Set Y-axis number format.

    Args:
        chart: An openpyxl chart object.
        fmt: Number format string (e.g. '#,##0', '0.0').
    """
    chart.y_axis.numFmt = fmt


def apply_axis_labels(
    chart: object,
    x_title: str | None = None,
    y_title: str | None = None,
    y_numfmt: str | None = None,
    x_numfmt: str | None = None,
) -> None:
    """Set axis titles and number formats on a chart.

    Args:
        chart: An openpyxl chart object.
        x_title: Optional X-axis title.
        y_title: Optional Y-axis title.
        y_numfmt: Optional Y-axis number format string.
        x_numfmt: Optional X-axis number format string.
    """
    if x_title is not None:
        chart.x_axis.title = x_title
    if y_title is not None:
        chart.y_axis.title = y_title
    if y_numfmt is not None:
        chart.y_axis.numFmt = y_numfmt
    if x_numfmt is not None:
        chart.x_axis.numFmt = x_numfmt


def apply_y_axis_padding(
    chart: object,
    data_min: float,
    data_max: float,
    padding: float = 0.10,
    divisions: int = 10,
) -> None:
    """Set Y-axis scaling with padding and tick interval.

    Args:
        chart: An openpyxl chart object.
        data_min: Minimum data value in the series.
        data_max: Maximum data value in the series.
        padding: Fraction of the range to add as padding (default 0.10 = 10%).
        divisions: Number of major tick divisions (default 10).
    """
    span = data_max - data_min
    if span == 0:
        span = abs(data_max) * 0.1 or 1.0
    axis_min = data_min - span * padding
    axis_max = data_max + span * padding
    chart.y_axis.scaling.min = axis_min
    chart.y_axis.scaling.max = axis_max
    chart.y_axis.majorUnit = (axis_max - axis_min) / divisions
    chart.y_axis.tickLblPos = "nextTo"
    chart.y_axis.delete = False


def apply_x_axis_tick_interval(
    chart: object,
    data_count: int,
    divisions: int = 10,
) -> None:
    """Set X-axis tick label skip so that roughly *divisions* labels appear.

    Args:
        chart: An openpyxl chart object.
        data_count: Total number of data points on the X-axis.
        divisions: Desired number of visible labels (default 10).
    """
    skip = max(1, data_count // divisions)
    chart.x_axis.tickLblSkip = skip
    chart.x_axis.tickMarkSkip = skip
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.delete = False


def add_right_y_axis(
    chart: object,
    ws: Worksheet,
    data_col: int,
    header_row: int,
    data_end_row: int,
    y_min: float | None = None,
    y_max: float | None = None,
    y_numfmt: str | None = None,
) -> None:
    """Add a right-side Y-axis that mirrors the primary axis scale.

    Creates an invisible secondary series so that openpyxl renders a
    secondary Y-axis on the right with matching scale and number format.

    Args:
        chart: Primary chart object.
        ws: Worksheet that holds the data.
        data_col: Column number of an existing data series to reference.
        header_row: Row containing the column header.
        data_end_row: Last data row.
        y_min: Y-axis minimum (should match primary).
        y_max: Y-axis maximum (should match primary).
        y_numfmt: Y-axis number format string.
    """
    from openpyxl.chart import LineChart as _LC
    from openpyxl.chart import Reference as _Ref
    from openpyxl.chart.legend import LegendEntry

    c2 = _LC()
    # Explicit axis IDs — required for openpyxl to serialize secondary axes
    c2.y_axis.axId = 200
    c2.x_axis.axId = 210
    c2.y_axis.crossAx = 210
    c2.x_axis.crossAx = 200

    ref = _Ref(ws, min_col=data_col, min_row=header_row, max_row=data_end_row)
    c2.add_data(ref, titles_from_data=True)

    # Make the secondary series invisible
    c2.series[0].graphicalProperties.line.noFill = True

    # Hide secondary X-axis so it doesn't overlay the primary date labels
    c2.x_axis.delete = True

    # Show secondary Y-axis labels on the right side
    c2.y_axis.delete = False
    c2.y_axis.tickLblPos = "nextTo"

    # Match primary Y-axis scaling and tick interval
    if y_min is not None:
        c2.y_axis.scaling.min = y_min
    if y_max is not None:
        c2.y_axis.scaling.max = y_max
    if y_min is not None and y_max is not None:
        c2.y_axis.majorUnit = (y_max - y_min) / 10
    if y_numfmt is not None:
        c2.y_axis.numFmt = y_numfmt

    hidden_idx = len(chart.series)
    chart += c2

    # Hide the invisible series from the legend
    if chart.legend is not None:
        entry = LegendEntry(idx=hidden_idx, delete=True)
        if chart.legend.legendEntry is None:
            chart.legend.legendEntry = []
        chart.legend.legendEntry.append(entry)


def style_line_series(
    series: object,
    color: str,
    width: int,
    dash: str | None = None,
) -> None:
    """Apply color, width, and optional dash style to a chart series line.

    Args:
        series: Chart series object.
        color: Hex color string (e.g. '2F5496').
        width: Line width in EMU.
        dash: Optional dash style ('dash', 'dot', etc.).
    """
    series.graphicalProperties.line.solidFill = color
    series.graphicalProperties.line.width = width
    if dash:
        series.graphicalProperties.line.dashStyle = dash


def write_constant_column(
    ws: Worksheet,
    col: int,
    start_row: int,
    count: int,
    value: float,
    header: str,
) -> None:
    """Write a constant value column for reference lines in charts.

    Args:
        ws: Target worksheet.
        col: Column number (1-based).
        start_row: Header row.
        count: Number of data rows (excluding header).
        value: Constant value to fill.
        header: Column header text.
    """
    from openpyxl.styles import Alignment as _Align

    style_cell(ws, start_row, col, header, font=HDR_FONT, fill=HDR_FILL,
               alignment=_Align(horizontal="center"))
    for i in range(1, count + 1):
        ws.cell(row=start_row + i, column=col, value=value)


def add_reference_line_series(
    chart: object,
    ws: Worksheet,
    col: int,
    header_row: int,
    data_end_row: int,
    color: str,
    label: str | None = None,
    width: int = LINE_WIDTH_REFERENCE,
) -> None:
    """Add a dashed reference line series to a chart from a constant column.

    Args:
        chart: Target chart object.
        ws: Worksheet containing the data.
        col: Column number with constant data.
        header_row: Row containing the header.
        data_end_row: Last data row.
        color: Hex color for the line.
        label: Optional legend label override.
        width: Line width in EMU.
    """
    from openpyxl.chart import Reference as _Ref

    ref = _Ref(ws, min_col=col, min_row=header_row, max_row=data_end_row)
    chart.add_data(ref, titles_from_data=True)
    idx = len(chart.series) - 1
    s = chart.series[idx]
    s.graphicalProperties.line.solidFill = color
    s.graphicalProperties.line.width = width
    s.graphicalProperties.line.dashStyle = "dash"
