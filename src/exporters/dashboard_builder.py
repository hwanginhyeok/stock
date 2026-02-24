"""Overview dashboard sheet builder for the Market Overview Excel.

Layout:
    Rows 1-2:   Title / Subtitle
    Row 4:      Normalized performance chart section header
    Rows 5-34:  Normalized performance chart (anchor A5)
    Row 36:     Individual index mini-charts section header
    Rows 37-53: Mini chart row 1 (3 charts)
    Rows 54-70: Mini chart row 2 (3 charts)
    Row 72:     Index summary table section header
    Rows 73+:   Summary table + VIX row

    Chart data is written in columns N+ (14+) to avoid overlap with
    the visible table area (A:L).
"""

from __future__ import annotations

from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from src.exporters.base import (
    DARK_BLUE,
    GRAY,
    HDR_FILL,
    HDR_FONT,
    LIGHT_BLUE,
    LINE_WIDTH_HEAVY,
    LINE_WIDTH_MEDIUM,
    LINK_FONT,
    SECTION_FILL,
    SECTION_FONT,
    THIN_BORDER,
    apply_axis_labels,
    apply_chart_gridlines,
    apply_x_axis_tick_interval,
    apply_y_axis_padding,
    pct_color_font,
    style_cell,
    style_line_series,
)

# Sheet name mapping for index tickers
INDEX_SHEET_NAMES: dict[str, str] = {
    "^GSPC": "SP500",
    "^IXIC": "NASDAQ",
    "^RUT": "Russell2000",
    "^DJI": "DOW",
    "KOSPI": "KOSPI",
    "KOSDAQ": "KOSDAQ",
}

# Chart data zone — right-side columns to avoid visible overlap
_DATA_COL = 14  # Column N

# Row layout constants
_NORM_SECTION_ROW = 4
_NORM_CHART_ROW = 5
_MINI_SECTION_ROW = 36
_MINI_CHART_START = 37
_MINI_ROW_HEIGHT = 17
_CHARTS_END_ROW = _MINI_SECTION_ROW + 1 + 2 * _MINI_ROW_HEIGHT  # 71

# Chart colors
_COLORS = ["2F5496", "ED7D31", "70AD47", "FFC000", "7030A0", "44546A"]

# Mini-chart grid configuration
_GRID_ORDER = ["^GSPC", "^IXIC", "^RUT", "^DJI", "KOSPI", "KOSDAQ"]
_GRID_NAMES: dict[str, str] = {
    "^GSPC": "S&P 500", "^IXIC": "NASDAQ", "^RUT": "Russell 2000",
    "^DJI": "DOW", "KOSPI": "KOSPI", "KOSDAQ": "KOSDAQ",
}


def _compute_period_return(df: pd.DataFrame, days: int) -> float | None:
    """Compute percentage return over a given number of trading days.

    Args:
        df: OHLCV DataFrame with Close column.
        days: Approximate trading days to look back.

    Returns:
        Percentage return or None if insufficient data.
    """
    if len(df) < days + 1 or "Close" not in df.columns:
        return None
    current = float(df["Close"].iloc[-1])
    past = float(df["Close"].iloc[-days - 1])
    if past == 0:
        return None
    return round((current / past - 1) * 100, 2)


def _compute_ytd_return(df: pd.DataFrame) -> float | None:
    """Compute YTD return from the first trading day of the year.

    Args:
        df: OHLCV DataFrame.

    Returns:
        YTD percentage return or None.
    """
    if df.empty or "Close" not in df.columns:
        return None
    current = float(df["Close"].iloc[-1])
    current_year = df.index[-1].year
    year_data = df[df.index.year == current_year]
    if year_data.empty:
        return None
    first_close = float(year_data["Close"].iloc[0])
    if first_close == 0:
        return None
    return round((current / first_close - 1) * 100, 2)


def _safe_date(dt: Any) -> Any:
    """Strip timezone from a datetime-like value for Excel compatibility."""
    if hasattr(dt, "tz_localize") and dt.tzinfo:
        return dt.tz_localize(None)
    return dt


def _write_chart_data(
    ws: Any,
    ordered_tickers: list[str],
    all_dfs: dict[str, pd.DataFrame],
    index_data: dict[str, dict[str, Any]],
    min_len: int,
    ref_dates: Any,
) -> tuple[int, int, int, int, float, float]:
    """Write normalized + mini-chart data to columns N+ (right side).

    Returns:
        Tuple of (norm_hdr_row, norm_end_row,
                  mini_hdr_row, mini_end_row,
                  norm_global_min, norm_global_max).
    """
    norm_hdr_row = _DATA_COL  # reuse as a constant; actually we start at row 4
    norm_hdr_row = 4

    # Normalized performance data header
    style_cell(ws, norm_hdr_row, _DATA_COL, "Date", font=HDR_FONT, fill=HDR_FILL,
               alignment=Alignment(horizontal="center"))
    for j, ticker in enumerate(ordered_tickers):
        if ticker in all_dfs:
            style_cell(ws, norm_hdr_row, _DATA_COL + 1 + j,
                       index_data[ticker]["name"],
                       font=HDR_FONT, fill=HDR_FILL,
                       alignment=Alignment(horizontal="center"))

    # Normalized data rows
    norm_global_min = float("inf")
    norm_global_max = float("-inf")
    for i in range(min_len):
        data_row = norm_hdr_row + 1 + i
        ws.cell(row=data_row, column=_DATA_COL,
                value=_safe_date(ref_dates[i])).number_format = "YYYY-MM-DD"

        for j, ticker in enumerate(ordered_tickers):
            if ticker not in all_dfs:
                continue
            close_series = all_dfs[ticker]["Close"].iloc[-min_len:]
            base = float(close_series.iloc[0])
            if base > 0:
                normalized = float(close_series.iloc[i]) / base * 100
                ws.cell(row=data_row, column=_DATA_COL + 1 + j,
                        value=round(normalized, 2))
                norm_global_min = min(norm_global_min, normalized)
                norm_global_max = max(norm_global_max, normalized)

    norm_end_row = norm_hdr_row + min_len

    # Mini chart data (below normalized data, same columns)
    mini_hdr_row = norm_end_row + 2

    style_cell(ws, mini_hdr_row, _DATA_COL, "Date", font=HDR_FONT, fill=HDR_FILL,
               alignment=Alignment(horizontal="center"))
    for i in range(min_len):
        ws.cell(row=mini_hdr_row + 1 + i, column=_DATA_COL,
                value=_safe_date(ref_dates[i])).number_format = "YYYY-MM-DD"

    for g_idx, ticker in enumerate(_GRID_ORDER):
        if ticker not in all_dfs:
            continue
        display = _GRID_NAMES.get(ticker, ticker)
        col_num = _DATA_COL + 1 + g_idx

        style_cell(ws, mini_hdr_row, col_num, display,
                   font=HDR_FONT, fill=HDR_FILL,
                   alignment=Alignment(horizontal="center"))

        close_series = all_dfs[ticker]["Close"].iloc[-min_len:]
        for i in range(min_len):
            ws.cell(row=mini_hdr_row + 1 + i, column=col_num,
                    value=round(float(close_series.iloc[i]), 2))

    mini_end_row = mini_hdr_row + min_len

    return (norm_hdr_row, norm_end_row,
            mini_hdr_row, mini_end_row,
            norm_global_min, norm_global_max)


def _add_normalized_chart(
    ws: Any,
    ordered_tickers: list[str],
    all_dfs: dict[str, pd.DataFrame],
    min_len: int,
    norm_hdr_row: int,
    norm_end_row: int,
    norm_global_min: float,
    norm_global_max: float,
) -> None:
    """Create and anchor the normalized performance chart at top of sheet."""
    # Section header
    ws.merge_cells(start_row=_NORM_SECTION_ROW, start_column=1,
                   end_row=_NORM_SECTION_ROW, end_column=12)
    style_cell(ws, _NORM_SECTION_ROW, 1,
               "Normalized Performance Comparison (Base=100)",
               font=SECTION_FONT, fill=SECTION_FILL)
    for c in range(2, 13):
        ws.cell(row=_NORM_SECTION_ROW, column=c).fill = SECTION_FILL
        ws.cell(row=_NORM_SECTION_ROW, column=c).border = THIN_BORDER

    chart = LineChart()
    chart.title = "Normalized Performance (Base = 100)"
    chart.style = 10
    chart.width = 32
    chart.height = 16
    chart.legend.position = "b"
    apply_chart_gridlines(chart)
    apply_axis_labels(chart, x_title="Date", y_title="Normalized Value",
                      y_numfmt="0.00", x_numfmt="yyyy-mm")
    if norm_global_min < float("inf"):
        apply_y_axis_padding(chart, norm_global_min, norm_global_max)
    apply_x_axis_tick_interval(chart, min_len)

    dates_ref = Reference(ws, min_col=_DATA_COL,
                          min_row=norm_hdr_row + 1, max_row=norm_end_row)

    active_cols = [j for j, t in enumerate(ordered_tickers) if t in all_dfs]
    for idx, j in enumerate(active_cols):
        data_ref = Reference(ws, min_col=_DATA_COL + 1 + j,
                             min_row=norm_hdr_row, max_row=norm_end_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(dates_ref)
        style_line_series(chart.series[idx],
                          _COLORS[idx % len(_COLORS)], LINE_WIDTH_HEAVY)

    ws.add_chart(chart, f"A{_NORM_CHART_ROW}")


def _add_mini_charts(
    ws: Any,
    all_dfs: dict[str, pd.DataFrame],
    min_len: int,
    mini_hdr_row: int,
    mini_end_row: int,
) -> None:
    """Create and anchor individual index mini-charts in a 3x2 grid."""
    # Section header
    ws.merge_cells(start_row=_MINI_SECTION_ROW, start_column=1,
                   end_row=_MINI_SECTION_ROW, end_column=12)
    style_cell(ws, _MINI_SECTION_ROW, 1, "Individual Index Trends",
               font=SECTION_FONT, fill=SECTION_FILL)
    for c in range(2, 13):
        ws.cell(row=_MINI_SECTION_ROW, column=c).fill = SECTION_FILL
        ws.cell(row=_MINI_SECTION_ROW, column=c).border = THIN_BORDER

    mini_date_ref = Reference(ws, min_col=_DATA_COL,
                              min_row=mini_hdr_row + 1, max_row=mini_end_row)

    for g_idx, ticker in enumerate(_GRID_ORDER):
        if ticker not in all_dfs:
            continue
        display = _GRID_NAMES.get(ticker, ticker)
        close_series = all_dfs[ticker]["Close"].iloc[-min_len:]

        mini = LineChart()
        mini.title = display
        mini.style = 10
        mini.width = 10.5
        mini.height = 9
        mini.legend = None
        apply_chart_gridlines(mini)
        apply_axis_labels(mini, y_numfmt="#,##0.00", x_numfmt="yy-mm")
        apply_y_axis_padding(mini, float(close_series.min()),
                             float(close_series.max()))
        apply_x_axis_tick_interval(mini, min_len)
        mini.x_axis.tickLblPos = "low"

        data_ref = Reference(ws, min_col=_DATA_COL + 1 + g_idx,
                             min_row=mini_hdr_row, max_row=mini_end_row)
        mini.add_data(data_ref, titles_from_data=True)
        mini.set_categories(mini_date_ref)
        style_line_series(mini.series[0],
                          _COLORS[g_idx % len(_COLORS)], LINE_WIDTH_MEDIUM)

        grid_col = g_idx % 3
        grid_row = g_idx // 3
        anchor_col_letters = ["A", "E", "I"]
        anchor_row = _MINI_CHART_START + grid_row * _MINI_ROW_HEIGHT
        ws.add_chart(mini, f"{anchor_col_letters[grid_col]}{anchor_row}")


def _write_summary_table(
    ws: Any,
    ordered_tickers: list[str],
    index_data: dict[str, dict[str, Any]],
    vix_ticker: str | None,
    start_row: int,
) -> None:
    """Write index summary table and VIX row below the chart zone."""
    # Section header
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=12)
    style_cell(ws, start_row, 1, "Index Summary",
               font=SECTION_FONT, fill=SECTION_FILL)
    for c in range(2, 13):
        ws.cell(row=start_row, column=c).fill = SECTION_FILL
        ws.cell(row=start_row, column=c).border = THIN_BORDER

    # Column headers
    hdr_row = start_row + 1
    headers = [
        ("No", 5), ("Market", 8), ("Index", 16), ("Current", 14),
        ("1D%", 8), ("1W%", 8), ("1M%", 8), ("3M%", 8),
        ("6M%", 8), ("YTD%", 8), ("Trend", 10), ("Detail", 8),
    ]
    for col_idx, (header, width) in enumerate(headers, 1):
        style_cell(ws, hdr_row, col_idx, header, font=HDR_FONT, fill=HDR_FILL,
                   alignment=Alignment(horizontal="center"))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    row = hdr_row + 1
    for i, ticker in enumerate(ordered_tickers):
        info = index_data[ticker]
        df = info["ohlcv"]
        current = info.get("current", 0)
        name = info["name"]
        market = info["market"]

        ret_1d = _compute_period_return(df, 1)
        ret_1w = _compute_period_return(df, 5)
        ret_1m = _compute_period_return(df, 21)
        ret_3m = _compute_period_return(df, 63)
        ret_6m = _compute_period_return(df, 126)
        ret_ytd = _compute_ytd_return(df)

        if ret_1m is not None and ret_3m is not None:
            if ret_1m > 0 and ret_3m > 0:
                trend = "Uptrend"
            elif ret_1m < 0 and ret_3m < 0:
                trend = "Downtrend"
            else:
                trend = "Mixed"
        else:
            trend = "N/A"

        style_cell(ws, row, 1, i + 1, alignment=Alignment(horizontal="center"))
        style_cell(ws, row, 2, market, alignment=Alignment(horizontal="center"))
        style_cell(ws, row, 3, name, font=Font(bold=True, size=10))
        style_cell(ws, row, 4, current, num_fmt="#,##0.00")

        for col_idx, ret_val in [(5, ret_1d), (6, ret_1w), (7, ret_1m),
                                  (8, ret_3m), (9, ret_6m), (10, ret_ytd)]:
            if ret_val is not None:
                style_cell(ws, row, col_idx, ret_val / 100,
                           font=pct_color_font(ret_val),
                           num_fmt="0.00%",
                           alignment=Alignment(horizontal="center"))
            else:
                style_cell(ws, row, col_idx, "N/A",
                           alignment=Alignment(horizontal="center"))

        trend_color = ("548235" if trend == "Uptrend"
                       else ("C00000" if trend == "Downtrend" else GRAY))
        style_cell(ws, row, 11, trend,
                   font=Font(bold=True, color=trend_color, size=10),
                   alignment=Alignment(horizontal="center"))

        sheet_name = INDEX_SHEET_NAMES.get(ticker, ticker.replace("^", ""))
        link_cell = style_cell(ws, row, 12, "Go >>", font=LINK_FONT,
                               alignment=Alignment(horizontal="center"))
        link_cell.hyperlink = f"#'{sheet_name}'!A1"
        row += 1

    # VIX row (special)
    if vix_ticker and vix_ticker in index_data:
        vix_info = index_data[vix_ticker]
        vix_current = vix_info.get("current", 0)
        vix_df = vix_info["ohlcv"]

        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        style_cell(ws, row, 1, "VIX", font=SECTION_FONT, fill=SECTION_FILL)
        ws.cell(row=row, column=2).fill = SECTION_FILL
        ws.cell(row=row, column=2).border = THIN_BORDER
        style_cell(ws, row, 3, "Fear & Greed Index",
                   font=SECTION_FONT, fill=SECTION_FILL)
        for c in range(4, 13):
            ws.cell(row=row, column=c).fill = SECTION_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

        style_cell(ws, row, 1, "", alignment=Alignment(horizontal="center"))
        style_cell(ws, row, 2, "US", alignment=Alignment(horizontal="center"))
        style_cell(ws, row, 3, "VIX (Volatility Index)",
                   font=Font(bold=True, size=10))
        style_cell(ws, row, 4, vix_current, num_fmt="#,##0.00")

        vix_ret_1d = _compute_period_return(vix_df, 1)
        if vix_ret_1d is not None:
            style_cell(ws, row, 5, vix_ret_1d / 100,
                       font=pct_color_font(-vix_ret_1d),
                       num_fmt="0.00%",
                       alignment=Alignment(horizontal="center"))

        if vix_current >= 30:
            vix_label = "High Fear"
        elif vix_current >= 20:
            vix_label = "Elevated"
        elif vix_current <= 12:
            vix_label = "Complacent"
        else:
            vix_label = "Normal"
        style_cell(ws, row, 11, vix_label,
                   font=Font(bold=True, size=10),
                   alignment=Alignment(horizontal="center"))


def build_overview_sheet(
    wb: Workbook,
    index_data: dict[str, dict[str, Any]],
    period_label: str,
) -> None:
    """Build the Overview sheet with charts at top and summary table below.

    Args:
        wb: Workbook instance (active sheet will be used).
        index_data: Dict keyed by ticker with keys:
            - name: display name
            - market: "US" or "KR"
            - ohlcv: DataFrame
            - current: current price/value
        period_label: Display label for the analysis period.
    """
    ws = wb.active
    ws.title = "Overview"

    # ===== Title (rows 1-2) =====
    ws.merge_cells("A1:L1")
    title_cell = ws.cell(row=1, column=1, value="Market Overview Dashboard")
    title_cell.font = Font(bold=True, size=16, color=DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:L2")
    date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.cell(
        row=2, column=1,
        value=f"Generated: {date_str}  |  Chart Period: {period_label}",
    ).font = Font(size=10, color=GRAY)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Prepare index lists
    ordered_tickers = [t for t in index_data if t != "^VIX"]
    vix_ticker = "^VIX" if "^VIX" in index_data else None

    all_dfs = {t: index_data[t]["ohlcv"] for t in ordered_tickers
               if not index_data[t]["ohlcv"].empty}

    table_start_row = _CHARTS_END_ROW + 1  # default

    if all_dfs:
        min_len = min(len(df) for df in all_dfs.values())
        ref_ticker = list(all_dfs.keys())[0]
        ref_dates = all_dfs[ref_ticker].index[-min_len:]

        # 1) Write chart data to right-side columns (N+)
        (norm_hdr_row, norm_end_row,
         mini_hdr_row, mini_end_row,
         norm_global_min, norm_global_max) = _write_chart_data(
            ws, ordered_tickers, all_dfs, index_data, min_len, ref_dates)

        # 2) Create normalized performance chart (rows 4-35)
        _add_normalized_chart(
            ws, ordered_tickers, all_dfs, min_len,
            norm_hdr_row, norm_end_row, norm_global_min, norm_global_max)

        # 3) Create mini charts (rows 36-70)
        _add_mini_charts(ws, all_dfs, min_len, mini_hdr_row, mini_end_row)

    # 4) Write summary table below chart zone
    _write_summary_table(ws, ordered_tickers, index_data, vix_ticker,
                         table_start_row)

    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = DARK_BLUE
