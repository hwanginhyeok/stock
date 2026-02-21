"""Sentiment sheet builder — VIX, Fear/Greed, trend strength, market diagnosis."""

from __future__ import annotations

from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from src.exporters.base import (
    DARK_BLUE,
    GRAY,
    GREEN,
    HDR_FILL,
    HDR_FONT,
    LABEL_FONT,
    LIGHT_GRAY,
    LINE_WIDTH_HEAVY,
    LINK_FONT,
    MED_BLUE,
    ORANGE,
    RED,
    SECTION_FILL,
    SECTION_FONT,
    THIN_BORDER,
    TITLE_FONT,
    VALUE_FONT,
    add_reference_line_series,
    apply_axis_labels,
    apply_chart_gridlines,
    apply_x_axis_tick_interval,
    apply_y_axis_padding,
    style_cell,
    style_line_series,
    write_constant_column,
    write_label_value,
    write_section_header,
)


def build_sentiment_sheet(
    wb: Workbook,
    vix_data: dict[str, Any],
    fear_greed: dict[str, Any],
    trend_data: list[dict[str, Any]],
    diagnosis: dict[str, Any],
) -> None:
    """Build the Sentiment sheet with VIX, Fear/Greed, trend, and diagnosis.

    Args:
        wb: Target Workbook.
        vix_data: Dict with "current", "ohlcv" keys for VIX.
        fear_greed: Fear/Greed index result from compute_fear_greed().
        trend_data: List of trend strength dicts.
        diagnosis: Market diagnosis dict.
    """
    ws = wb.create_sheet(title="Sentiment")

    # Title
    ws.merge_cells("A1:K1")
    ws.cell(row=1, column=1, value="Market Sentiment & Trend Analysis").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    # Back link
    back_cell = ws.cell(row=1, column=12, value="<< Overview")
    back_cell.font = LINK_FONT
    back_cell.hyperlink = "#Overview!A1"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 3

    row = 3

    # ===== VIX Section =====
    row = write_section_header(ws, row, "VIX (Volatility Index)")

    vix_current = vix_data.get("current", 0)
    row = write_label_value(ws, row, "Current VIX", f"{vix_current:.2f}",
                            val_font=Font(bold=True, size=14))

    # VIX level interpretation
    if vix_current >= 30:
        vix_level = "High Volatility (Extreme Fear)"
        vix_color = RED
    elif vix_current >= 20:
        vix_level = "Elevated Volatility (Caution)"
        vix_color = ORANGE
    elif vix_current <= 12:
        vix_level = "Low Volatility (Complacency)"
        vix_color = GREEN
    else:
        vix_level = "Normal Volatility"
        vix_color = MED_BLUE

    row = write_label_value(ws, row, "Level", vix_level,
                            val_font=Font(bold=True, color=vix_color, size=11))

    # VIX thresholds reference
    row += 1
    row = write_label_value(ws, row, "< 12", "Low (Complacent)")
    row = write_label_value(ws, row, "12 ~ 20", "Normal")
    row = write_label_value(ws, row, "20 ~ 30", "Elevated (Caution)")
    row = write_label_value(ws, row, "> 30", "High (Fear)")
    row += 1

    # VIX chart data + chart
    vix_ohlcv = vix_data.get("ohlcv")
    if vix_ohlcv is not None and not vix_ohlcv.empty and "Close" in vix_ohlcv.columns:
        chart_data_start = row
        style_cell(ws, chart_data_start, 4, "Date", font=HDR_FONT, fill=HDR_FILL,
                   alignment=Alignment(horizontal="center"))
        style_cell(ws, chart_data_start, 5, "VIX", font=HDR_FONT, fill=HDR_FILL,
                   alignment=Alignment(horizontal="center"))
        ws.column_dimensions["D"].width = 13
        ws.column_dimensions["E"].width = 13

        vix_n = len(vix_ohlcv)
        for i in range(vix_n):
            r = chart_data_start + 1 + i
            dt = vix_ohlcv.index[i]
            date_val = dt.tz_localize(None) if hasattr(dt, 'tz_localize') and dt.tzinfo else dt
            ws.cell(row=r, column=4, value=date_val).number_format = "YYYY-MM-DD"
            ws.cell(row=r, column=5, value=round(float(vix_ohlcv["Close"].iloc[i]), 2))

        data_end = chart_data_start + vix_n

        # VIX reference-line constant columns (F=6, G=7, H=8)
        write_constant_column(ws, 6, chart_data_start, vix_n, 12, "VIX 12")
        write_constant_column(ws, 7, chart_data_start, vix_n, 20, "VIX 20")
        write_constant_column(ws, 8, chart_data_start, vix_n, 30, "VIX 30")
        for col_letter in ("F", "G", "H"):
            ws.column_dimensions[col_letter].width = 10

        vix_chart = LineChart()
        vix_chart.title = "VIX History"
        vix_chart.style = 10
        vix_chart.width = 32
        vix_chart.height = 14
        vix_chart.legend.position = "b"
        apply_chart_gridlines(vix_chart)
        apply_axis_labels(vix_chart, x_title="Date", y_title="VIX",
                          y_numfmt="0.00", x_numfmt="yyyy-mm")
        vix_close = vix_ohlcv["Close"]
        vix_data_min = min(float(vix_close.min()), 12)  # include ref lines
        vix_data_max = max(float(vix_close.max()), 30)
        apply_y_axis_padding(vix_chart, vix_data_min, vix_data_max)
        apply_x_axis_tick_interval(vix_chart, vix_n)

        dates_ref = Reference(ws, min_col=4, min_row=chart_data_start + 1,
                              max_row=data_end)
        vix_ref = Reference(ws, min_col=5, min_row=chart_data_start,
                            max_row=data_end)
        vix_chart.add_data(vix_ref, titles_from_data=True)
        vix_chart.set_categories(dates_ref)
        style_line_series(vix_chart.series[0], "C00000", LINE_WIDTH_HEAVY)

        # VIX 12 (green), 20 (orange), 30 (red) reference lines
        add_reference_line_series(vix_chart, ws, 6, chart_data_start, data_end, GREEN)
        add_reference_line_series(vix_chart, ws, 7, chart_data_start, data_end, ORANGE)
        add_reference_line_series(vix_chart, ws, 8, chart_data_start, data_end, RED)

        ws.add_chart(vix_chart, "D1")

        row = max(row, data_end + 2)
    else:
        row += 1

    # ===== Fear & Greed Section =====
    row += 1
    row = write_section_header(ws, row, "Fear & Greed Index (Custom)")

    fg_score = fear_greed["score"]
    fg_level = fear_greed["level"]

    # Color based on level
    if fg_score >= 60:
        fg_color = GREEN
    elif fg_score <= 40:
        fg_color = RED
    else:
        fg_color = ORANGE

    row = write_label_value(ws, row, "Score", f"{fg_score:.1f} / 100",
                            val_font=Font(bold=True, size=14, color=fg_color))
    row = write_label_value(ws, row, "Level", fg_level,
                            val_font=Font(bold=True, color=fg_color, size=11))
    row += 1

    # Components breakdown
    components = fear_greed.get("components", {})
    row = write_label_value(ws, row, "Component", "Score (Weight)")
    vix_comp = components.get("vix", {})
    row = write_label_value(ws, row, f"VIX ({vix_comp.get('value', 'N/A')})",
                            f"{vix_comp.get('score', 'N/A')} (40%)")
    rsi_comp = components.get("rsi", {})
    row = write_label_value(ws, row, f"Avg RSI ({rsi_comp.get('value', 'N/A')})",
                            f"{rsi_comp.get('score', 'N/A')} (30%)")
    momentum_comp = components.get("momentum", {})
    row = write_label_value(ws, row, f"Momentum ({momentum_comp.get('value', 'N/A')}%)",
                            f"{momentum_comp.get('score', 'N/A')} (30%)")
    row += 2

    # ===== Trend Strength Table =====
    row = write_section_header(ws, row, "Trend Strength Summary", cols=8)

    # Headers
    trend_headers = [
        ("Index", 16), ("ADX", 8), ("ADX Trend", 12),
        ("Supertrend", 12), ("RSI", 8), ("RSI State", 12),
    ]
    for col_idx, (header, width) in enumerate(trend_headers, 1):
        style_cell(ws, row, col_idx, header, font=HDR_FONT, fill=HDR_FILL,
                   alignment=Alignment(horizontal="center"))
    row += 1

    for t in trend_data:
        style_cell(ws, row, 1, t["name"], font=Font(bold=True, size=10))
        style_cell(ws, row, 2, t.get("adx", "N/A"),
                   num_fmt="0.0" if t.get("adx") is not None else None,
                   alignment=Alignment(horizontal="center"))

        adx_trend = t.get("adx_trend", "N/A")
        adx_color = GREEN if adx_trend == "Strong" else (ORANGE if adx_trend == "Moderate" else GRAY)
        style_cell(ws, row, 3, adx_trend,
                   font=Font(color=adx_color, bold=True, size=10),
                   alignment=Alignment(horizontal="center"))

        st_dir = t.get("supertrend_direction", "N/A")
        st_color = GREEN if st_dir == "Bullish" else (RED if st_dir == "Bearish" else GRAY)
        style_cell(ws, row, 4, st_dir,
                   font=Font(color=st_color, bold=True, size=10),
                   alignment=Alignment(horizontal="center"))

        style_cell(ws, row, 5, t.get("rsi", "N/A"),
                   num_fmt="0.0" if t.get("rsi") is not None else None,
                   alignment=Alignment(horizontal="center"))

        rsi_state = t.get("rsi_state", "N/A")
        rsi_color = RED if rsi_state == "Overbought" else (GREEN if rsi_state == "Oversold" else GRAY)
        style_cell(ws, row, 6, rsi_state,
                   font=Font(color=rsi_color, size=10),
                   alignment=Alignment(horizontal="center"))
        row += 1

    row += 2

    # ===== Market Diagnosis =====
    row = write_section_header(ws, row, "Market Diagnosis", cols=8)

    verdict = diagnosis.get("verdict", "N/A")
    if "Bullish" in verdict:
        verdict_color = GREEN
    elif "Bearish" in verdict:
        verdict_color = RED
    else:
        verdict_color = ORANGE

    row = write_label_value(ws, row, "Verdict", verdict,
                            val_font=Font(bold=True, size=14, color=verdict_color))

    description = diagnosis.get("description", "")
    # Write description spanning multiple columns
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=8)
    desc_cell = ws.cell(row=row, column=1, value=description)
    desc_cell.font = Font(size=10)
    desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
    row += 4

    row = write_label_value(ws, row, "Fear/Greed Score",
                            f"{diagnosis.get('fear_greed_score', 'N/A')}")
    row = write_label_value(ws, row, "Bullish Signals",
                            f"{diagnosis.get('bullish_signals', 0)}")
    row = write_label_value(ws, row, "Bearish Signals",
                            f"{diagnosis.get('bearish_signals', 0)}")

    ws.sheet_properties.tabColor = "C00000"
