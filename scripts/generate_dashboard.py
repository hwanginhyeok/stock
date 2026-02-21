"""Generate Market Overview Dashboard Excel.

Usage:
    python scripts/generate_dashboard.py                  # Default 6 months
    python scripts/generate_dashboard.py --period 1mo     # 1 month
    python scripts/generate_dashboard.py --period 3mo     # 3 months
    python scripts/generate_dashboard.py --period 1y      # 1 year
    python scripts/generate_dashboard.py --no-cache       # Ignore cache
    python scripts/generate_dashboard.py --top-n 15       # Top 15 stocks
"""

from __future__ import annotations

import argparse
import os
import shutil
import time
from datetime import datetime
from pathlib import Path
from typing import Any

# curl_cffi (used by yfinance) cannot load SSL certificates from non-ASCII
# paths. If the project directory contains non-ASCII characters, copy the
# certifi CA bundle to a temp location with an ASCII-only path.
import certifi as _certifi

_cert_src = Path(_certifi.where())
try:
    str(_cert_src).encode("ascii")
except (UnicodeEncodeError, UnicodeDecodeError):
    _cert_dst = Path(os.environ.get("TEMP", "/tmp")) / "yf_certs" / "cacert.pem"
    if not _cert_dst.exists() or _cert_dst.stat().st_size != _cert_src.stat().st_size:
        _cert_dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(_cert_src, _cert_dst)
    os.environ.setdefault("CURL_CA_BUNDLE", str(_cert_dst))

import pandas as pd
import yfinance as yf
from openpyxl import Workbook

from src.analyzers.market_sentiment import (
    compute_fear_greed,
    compute_trend_strength,
    market_diagnosis,
)
from src.analyzers.technical import TechnicalAnalyzer
from src.collectors.market.korea_collector import KoreaMarketCollector
from src.collectors.market.us_collector import USMarketCollector
from src.core.cache_manager import CacheManager
from src.core.config import PROJECT_ROOT, get_config
from src.exporters.constituents_builder import build_constituents_sheet
from src.exporters.dashboard_builder import INDEX_SHEET_NAMES, build_overview_sheet
from src.exporters.index_detail_builder import build_index_detail_sheet
from src.exporters.sentiment_builder import build_sentiment_sheet

# ============================================================
# Constants
# ============================================================

PERIOD_TO_DAYS: dict[str, int] = {
    "1mo": 30,
    "3mo": 90,
    "6mo": 180,
    "1y": 365,
    "2y": 730,
}

PERIOD_LABELS: dict[str, str] = {
    "1mo": "1 Month",
    "3mo": "3 Months",
    "6mo": "6 Months",
    "1y": "1 Year",
    "2y": "2 Years",
}

# Index tickers to collect (non-VIX)
US_INDEX_TICKERS = [
    ("^GSPC", "S&P 500"),
    ("^IXIC", "NASDAQ"),
    ("^RUT", "Russell 2000"),
    ("^DJI", "DOW"),
]

VIX_TICKER = ("^VIX", "VIX")

KR_INDEX_TICKERS = [
    ("KOSPI", "KOSPI"),
    ("KOSDAQ", "KOSDAQ"),
]

OUTPUT_DIR = PROJECT_ROOT / "data" / "processed" / "dashboards"


# ============================================================
# Data collection
# ============================================================


def _collect_us_index_ohlcv(
    ticker: str,
    period: str,
    cache: CacheManager,
) -> pd.DataFrame:
    """Collect OHLCV for a US index/stock via yfinance with caching.

    Args:
        ticker: yfinance ticker symbol.
        period: yfinance period string.
        cache: CacheManager instance.

    Returns:
        OHLCV DataFrame.
    """
    cache_key = f"{ticker.replace('^', '')}_{period}"
    cached = cache.get("indices", cache_key)
    if cached is not None:
        return cached

    print(f"  Fetching {ticker} ({period})...", end=" ", flush=True)
    try:
        yf_ticker = yf.Ticker(ticker)
        df = yf_ticker.history(period=period)
        if df.empty:
            print("EMPTY")
            return pd.DataFrame()
        standard_cols = ["Open", "High", "Low", "Close", "Volume"]
        available = [c for c in standard_cols if c in df.columns]
        df = df[available]
        cache.put("indices", cache_key, df)
        print(f"OK ({len(df)} rows)")
        return df
    except Exception as e:
        print(f"FAILED ({e})")
        return pd.DataFrame()


def _collect_kr_index_ohlcv(
    code: str,
    name: str,
    days: int,
    cache: CacheManager,
) -> pd.DataFrame:
    """Collect OHLCV for a Korean index via pykrx with caching.

    Args:
        code: pykrx index code.
        name: Display name.
        days: Number of calendar days.
        cache: CacheManager instance.

    Returns:
        OHLCV DataFrame.
    """
    cache_key = f"{name}_{days}d"
    cached = cache.get("indices", cache_key)
    if cached is not None:
        return cached

    print(f"  Fetching {name} ({days}d)...", end=" ", flush=True)
    try:
        from pykrx import stock as pykrx_stock
        today = datetime.now()
        start_str = (today - pd.Timedelta(days=days)).strftime("%Y%m%d")
        end_str = today.strftime("%Y%m%d")
        df = pykrx_stock.get_index_ohlcv(start_str, end_str, code)
        if df.empty:
            print("EMPTY")
            return pd.DataFrame()
        column_map = {
            "시가": "Open", "고가": "High", "저가": "Low",
            "종가": "Close", "거래량": "Volume",
        }
        df = df.rename(columns=column_map)
        standard_cols = ["Open", "High", "Low", "Close", "Volume"]
        available = [c for c in standard_cols if c in df.columns]
        df = df[available]
        cache.put("indices", cache_key, df)
        print(f"OK ({len(df)} rows)")
        return df
    except Exception as e:
        print(f"FAILED ({e})")
        return pd.DataFrame()


def _collect_stock_ohlcv(
    ticker: str,
    period: str,
    cache: CacheManager,
    market: str = "US",
) -> pd.DataFrame:
    """Collect OHLCV for an individual stock with caching.

    Args:
        ticker: Stock ticker.
        period: yfinance period string (for US) or days equiv.
        cache: CacheManager instance.
        market: "US" or "KR".

    Returns:
        OHLCV DataFrame.
    """
    cache_key = f"{ticker}_{period}"
    cached = cache.get("stocks", cache_key)
    if cached is not None:
        return cached

    try:
        if market == "US":
            yf_ticker = yf.Ticker(ticker)
            df = yf_ticker.history(period=period)
            if df.empty:
                return pd.DataFrame()
            standard_cols = ["Open", "High", "Low", "Close", "Volume"]
            available = [c for c in standard_cols if c in df.columns]
            df = df[available]
        else:
            from pykrx import stock as pykrx_stock
            days = PERIOD_TO_DAYS.get(period, 180)
            today = datetime.now()
            start_str = (today - pd.Timedelta(days=days)).strftime("%Y%m%d")
            end_str = today.strftime("%Y%m%d")
            df = pykrx_stock.get_market_ohlcv(start_str, end_str, ticker)
            if df.empty:
                return pd.DataFrame()
            column_map = {
                "시가": "Open", "고가": "High", "저가": "Low",
                "종가": "Close", "거래량": "Volume",
            }
            df = df.rename(columns=column_map)
            standard_cols = ["Open", "High", "Low", "Close", "Volume"]
            available = [c for c in standard_cols if c in df.columns]
            df = df[available]

        cache.put("stocks", cache_key, df)
        return df
    except Exception:
        return pd.DataFrame()


def _fetch_market_cap(ticker: str) -> float | None:
    """Fetch market cap from yfinance."""
    try:
        info = yf.Ticker(ticker).info
        return info.get("marketCap")
    except Exception:
        return None


# ============================================================
# Main pipeline
# ============================================================


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate Market Overview Dashboard")
    parser.add_argument("--period", default="6mo",
                        choices=["1mo", "3mo", "6mo", "1y", "2y"],
                        help="Chart data period (default: 6mo)")
    parser.add_argument("--no-cache", action="store_true",
                        help="Ignore cache and fetch all data fresh")
    parser.add_argument("--top-n", type=int, default=10,
                        help="Number of top stocks per market (default: 10)")
    args = parser.parse_args()

    period = args.period
    days = PERIOD_TO_DAYS[period]
    no_cache = args.no_cache
    top_n = args.top_n

    config = get_config()
    cache_config = config.market.cache
    index_cache = CacheManager(ttl_hours=cache_config.indices_ttl_hours, no_cache=no_cache)
    stock_cache = CacheManager(ttl_hours=cache_config.stocks_ttl_hours, no_cache=no_cache)

    print("=" * 60)
    print("  Market Overview Dashboard Generator")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"  Period: {PERIOD_LABELS[period]}  |  Cache: {'OFF' if no_cache else 'ON'}")
    print("=" * 60)
    print()

    # ===== Step 1: Collect Index Data =====
    print("[1/5] Collecting index data...")
    index_data: dict[str, dict[str, Any]] = {}

    # US indices
    for ticker, name in US_INDEX_TICKERS:
        df = _collect_us_index_ohlcv(ticker, period, index_cache)
        if not df.empty:
            current = float(df["Close"].iloc[-1])
            index_data[ticker] = {
                "name": name,
                "market": "US",
                "ohlcv": df,
                "current": current,
            }

    # VIX
    vix_df = _collect_us_index_ohlcv(VIX_TICKER[0], period, index_cache)
    vix_current = 0.0
    if not vix_df.empty:
        vix_current = float(vix_df["Close"].iloc[-1])
        index_data[VIX_TICKER[0]] = {
            "name": VIX_TICKER[1],
            "market": "US",
            "ohlcv": vix_df,
            "current": vix_current,
        }

    # Korean indices
    kr_index_codes = {
        "KOSPI": "1001",
        "KOSDAQ": "2001",
    }
    for name, code in kr_index_codes.items():
        df = _collect_kr_index_ohlcv(code, name, days, index_cache)
        if not df.empty:
            current = float(df["Close"].iloc[-1])
            index_data[name] = {
                "name": name,
                "market": "KR",
                "ohlcv": df,
                "current": current,
            }

    print(f"  Collected {len(index_data)} indices\n")

    # ===== Step 2: Collect Watchlist Stocks =====
    print("[2/5] Collecting watchlist stock data...")

    ta = TechnicalAnalyzer()

    # US watchlist
    us_watchlist = config.market.us.watchlist[:top_n]
    us_stocks: list[dict[str, Any]] = []
    for item in us_watchlist:
        print(f"  US: {item.ticker} ({item.name})...", end=" ", flush=True)
        df = _collect_stock_ohlcv(item.ticker, period, stock_cache, "US")
        if df.empty:
            print("SKIP")
            continue

        current = float(df["Close"].iloc[-1])
        change_1d = None
        change_1w = None
        if len(df) >= 2:
            change_1d = round((current / float(df["Close"].iloc[-2]) - 1) * 100, 2)
        if len(df) >= 6:
            change_1w = round((current / float(df["Close"].iloc[-6]) - 1) * 100, 2)

        tech_score = 50.0
        try:
            tech_result = ta.analyze(item.ticker, ohlcv=df)
            tech_score = tech_result["score"]
        except Exception:
            pass

        market_cap = _fetch_market_cap(item.ticker)

        us_stocks.append({
            "ticker": item.ticker,
            "name": item.name,
            "ohlcv": df,
            "current": current,
            "change_1d": change_1d,
            "change_1w": change_1w,
            "market_cap": market_cap,
            "tech_score": tech_score,
        })
        print(f"OK (Score={tech_score:.1f})")
        time.sleep(0.3)

    # KR watchlist
    kr_watchlist = config.market.korea.watchlist[:top_n]
    kr_stocks: list[dict[str, Any]] = []
    for item in kr_watchlist:
        print(f"  KR: {item.ticker} ({item.name})...", end=" ", flush=True)
        df = _collect_stock_ohlcv(item.ticker, period, stock_cache, "KR")
        if df.empty:
            print("SKIP")
            continue

        current = float(df["Close"].iloc[-1])
        change_1d = None
        change_1w = None
        if len(df) >= 2:
            change_1d = round((current / float(df["Close"].iloc[-2]) - 1) * 100, 2)
        if len(df) >= 6:
            change_1w = round((current / float(df["Close"].iloc[-6]) - 1) * 100, 2)

        tech_score = 50.0
        try:
            tech_result = ta.analyze(item.ticker, ohlcv=df)
            tech_score = tech_result["score"]
        except Exception:
            pass

        kr_stocks.append({
            "ticker": item.ticker,
            "name": item.name,
            "ohlcv": df,
            "current": current,
            "change_1d": change_1d,
            "change_1w": change_1w,
            "market_cap": None,
            "tech_score": tech_score,
        })
        print(f"OK (Score={tech_score:.1f})")
        time.sleep(0.3)

    print(f"  US: {len(us_stocks)} stocks, KR: {len(kr_stocks)} stocks\n")

    # ===== Step 3: Compute Sentiment =====
    print("[3/5] Computing sentiment indicators...")

    # Fear/Greed
    non_vix_data = {k: v["ohlcv"] for k, v in index_data.items() if k != "^VIX"}
    fear_greed = compute_fear_greed(vix_current, non_vix_data)
    print(f"  Fear/Greed: {fear_greed['score']:.1f} ({fear_greed['level']})")

    # Trend strength for each index
    trend_data: list[dict[str, Any]] = []
    for ticker, info in index_data.items():
        if ticker == "^VIX":
            continue
        ts = compute_trend_strength(ticker, info["name"], info["ohlcv"])
        trend_data.append(ts)
        print(f"  {info['name']}: ADX={ts.get('adx', 'N/A')}, "
              f"ST={ts.get('supertrend_direction', 'N/A')}, "
              f"RSI={ts.get('rsi', 'N/A')}")

    # Diagnosis
    diagnosis = market_diagnosis(fear_greed, trend_data)
    print(f"  Diagnosis: {diagnosis['verdict']}\n")

    # ===== Step 4: Build Excel =====
    print("[4/5] Building Excel workbook...")
    wb = Workbook()

    # Sheet 1: Overview
    build_overview_sheet(wb, index_data, PERIOD_LABELS[period])
    print("  [OK] Overview sheet")

    # Sheets 2-7: Individual index details
    for ticker, info in index_data.items():
        if ticker == "^VIX":
            continue
        sheet_name = INDEX_SHEET_NAMES.get(ticker, ticker.replace("^", ""))
        build_index_detail_sheet(wb, sheet_name, info["name"], info["ohlcv"])
        print(f"  [OK] {info['name']} detail sheet")

    # Sheet 8: Top_Stocks
    build_constituents_sheet(wb, us_stocks, kr_stocks)
    print("  [OK] Top_Stocks sheet")

    # Sheet 9: Sentiment
    vix_data_for_sheet = {
        "current": vix_current,
        "ohlcv": vix_df if not vix_df.empty else None,
    }
    build_sentiment_sheet(wb, vix_data_for_sheet, fear_greed, trend_data, diagnosis)
    print("  [OK] Sentiment sheet")

    # ===== Step 5: Save =====
    print("\n[5/5] Saving...")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now().strftime("%Y%m%d")
    output_path = OUTPUT_DIR / f"Market_Overview_{date_str}.xlsx"
    try:
        wb.save(str(output_path))
    except PermissionError:
        # File may be open in Excel; save with a version suffix instead
        for seq in range(2, 100):
            alt = OUTPUT_DIR / f"Market_Overview_{date_str}_v{seq}.xlsx"
            if not alt.exists():
                output_path = alt
                break
        wb.save(str(output_path))
        print(f"  (original file locked, saved as {output_path.name})")

    file_size = output_path.stat().st_size / 1024
    sheet_count = len(wb.sheetnames)

    print(f"\n{'=' * 60}")
    print(f"  Dashboard saved: {output_path}")
    print(f"  File size: {file_size:.0f} KB")
    print(f"  Sheets: {sheet_count} ({', '.join(wb.sheetnames)})")
    print(f"  Period: {PERIOD_LABELS[period]}")
    print(f"  Fear/Greed: {fear_greed['score']:.1f} ({fear_greed['level']})")
    print(f"  Diagnosis: {diagnosis['verdict']}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
