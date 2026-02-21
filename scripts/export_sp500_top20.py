"""S&P 500 Top 20 stocks: analysis report + charts in a single Excel."""

import sys
import time
from datetime import datetime
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

# Project imports
from src.analyzers.fundamental import FundamentalAnalyzer
from src.analyzers.screener import StockScreener
from src.analyzers.technical import (
    TechnicalAnalyzer,
    _bbands,
    _ema,
    _macd,
    _rsi,
    _sma,
)
from src.collectors.market.us_collector import USMarketCollector
from src.core.models import Market

# ============================================================
# Configuration
# ============================================================

TOP20 = [
    ("AAPL", "Apple"),
    ("MSFT", "Microsoft"),
    ("NVDA", "NVIDIA"),
    ("AMZN", "Amazon"),
    ("GOOGL", "Alphabet"),
    ("META", "Meta Platforms"),
    ("BRK-B", "Berkshire Hathaway"),
    ("TSLA", "Tesla"),
    ("AVGO", "Broadcom"),
    ("JPM", "JPMorgan Chase"),
    ("LLY", "Eli Lilly"),
    ("V", "Visa"),
    ("UNH", "UnitedHealth"),
    ("MA", "Mastercard"),
    ("XOM", "ExxonMobil"),
    ("COST", "Costco"),
    ("HD", "Home Depot"),
    ("PG", "Procter & Gamble"),
    ("JNJ", "Johnson & Johnson"),
    ("ABBV", "AbbVie"),
]

OUTPUT_PATH = Path("data/processed/stocks/SP500_Top20_Analysis.xlsx")

# Styles
DARK_BLUE = "2F5496"
MED_BLUE = "4472C4"
LIGHT_BLUE = "D6E4F0"
GREEN = "548235"
RED = "C00000"
ORANGE = "ED7D31"
GRAY = "808080"

HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
SECTION_FILL = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
LABEL_FONT = Font(bold=True, size=10)
VALUE_FONT = Font(size=10)
TITLE_FONT = Font(bold=True, size=14, color=DARK_BLUE)
THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
THIN_BOTTOM = Border(bottom=Side(style="thin"))


def style_cell(ws, row, col, value, font=None, fill=None, alignment=None, num_fmt=None):
    """Write and style a single cell."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if num_fmt:
        cell.number_format = num_fmt
    cell.border = THIN
    return cell


def rec_color(rec: str) -> str:
    """Return color hex for recommendation."""
    if rec == "strong_positive":
        return GREEN
    if rec == "positive":
        return "70AD47"
    if rec == "neutral":
        return ORANGE
    return RED


def score_fill(score: float) -> PatternFill:
    """Return fill color based on score."""
    if score >= 80:
        return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    if score >= 60:
        return PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    if score >= 40:
        return PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


# ============================================================
# Data collection & analysis
# ============================================================

def _format_market_cap(val: float | None) -> str:
    """Format market cap to human-readable string (e.g., '3.45T', '890B')."""
    if val is None or val == 0:
        return "N/A"
    if val >= 1e12:
        return f"${val / 1e12:.2f}T"
    if val >= 1e9:
        return f"${val / 1e9:.1f}B"
    if val >= 1e6:
        return f"${val / 1e6:.0f}M"
    return f"${val:,.0f}"


def _fetch_market_cap(ticker: str) -> float | None:
    """Fetch market cap from yfinance info."""
    try:
        import yfinance as yf
        info = yf.Ticker(ticker).info
        return info.get("marketCap")
    except Exception:
        return None


def collect_and_analyze():
    """Collect OHLCV and run analysis for all 20 stocks."""
    collector = USMarketCollector()
    ta = TechnicalAnalyzer()
    fa = FundamentalAnalyzer()
    screener = StockScreener()

    results = []
    for i, (ticker, name) in enumerate(TOP20):
        print(f"[{i+1:2d}/20] {ticker:6s} ({name})...", end=" ", flush=True)

        # Collect OHLCV
        ohlcv = collector.collect_stock_ohlcv(ticker, days=180)
        if ohlcv.empty:
            print("SKIP (no data)")
            continue

        # Technical
        tech = ta.analyze(ticker, ohlcv=ohlcv)
        # Fundamental (also contains market_cap in data)
        fund = fa.analyze(ticker, market=Market.US)
        # Composite
        comp = screener.analyze(ticker, ohlcv=ohlcv, market=Market.US)

        # Market cap from fundamental data or direct fetch
        market_cap = fund.get("data", {}).get("market_cap")
        if not market_cap:
            market_cap = _fetch_market_cap(ticker)

        results.append({
            "ticker": ticker,
            "name": name,
            "ohlcv": ohlcv,
            "tech": tech,
            "fund": fund,
            "comp": comp,
            "market_cap": market_cap,
        })
        mc_str = _format_market_cap(market_cap)
        print(f"MCap={mc_str}  Score={comp['score']:.1f} ({comp['recommendation']})")

        # Small delay to avoid rate limiting
        if i < len(TOP20) - 1:
            time.sleep(0.5)

    return results


# ============================================================
# Sheet 1: Summary list
# ============================================================

def build_summary_sheet(wb: Workbook, results: list[dict]):
    """Build the summary overview sheet."""
    ws = wb.active
    ws.title = "Overview"

    # Title
    ws.merge_cells("A1:L1")
    title_cell = ws.cell(row=1, column=1, value="S&P 500 Top 20 - Analysis Report")
    title_cell.font = Font(bold=True, size=16, color=DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Date
    ws.merge_cells("A2:L2")
    date_cell = ws.cell(row=2, column=1, value=f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    date_cell.font = Font(size=10, color=GRAY)
    date_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Headers
    headers = [
        ("No", 5), ("Ticker", 10), ("Company", 22),
        ("Market Cap", 16), ("Price (USD)", 14), ("Volume", 16),
        ("Tech Score", 12), ("Fund Score", 12), ("Composite", 12),
        ("Recommendation", 16), ("Signals", 40), ("Link", 8),
    ]
    for col_idx, (header, width) in enumerate(headers, 1):
        style_cell(ws, 4, col_idx, header, font=HDR_FONT, fill=HDR_FILL,
                   alignment=Alignment(horizontal="center"))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows — sorted by market cap descending
    results_sorted = sorted(
        results,
        key=lambda r: r.get("market_cap") or 0,
        reverse=True,
    )
    for i, r in enumerate(results_sorted):
        row = i + 5
        ohlcv = r["ohlcv"]
        comp = r["comp"]
        price = float(ohlcv["Close"].iloc[-1])
        volume = int(ohlcv["Volume"].iloc[-1])
        market_cap = r.get("market_cap")
        tech_score = comp["technical_score"]
        fund_score = comp["fundamental_score"]
        composite = comp["score"]
        rec = comp["recommendation"]
        signals = ", ".join(comp["signals"]) if comp["signals"] else "-"

        rec_display = {
            "strong_positive": "Strong Positive",
            "positive": "Positive",
            "neutral": "Neutral",
            "negative": "Negative",
        }.get(rec, rec)

        style_cell(ws, row, 1, i + 1, alignment=Alignment(horizontal="center"))
        style_cell(ws, row, 2, r["ticker"], font=Font(bold=True))
        style_cell(ws, row, 3, r["name"])
        style_cell(ws, row, 4, _format_market_cap(market_cap),
                   alignment=Alignment(horizontal="right"))
        style_cell(ws, row, 5, price, num_fmt="#,##0.00")
        style_cell(ws, row, 6, volume, num_fmt="#,##0")
        style_cell(ws, row, 7, tech_score, num_fmt="0.0",
                   fill=score_fill(tech_score))
        style_cell(ws, row, 8, fund_score, num_fmt="0.0",
                   fill=score_fill(fund_score))
        style_cell(ws, row, 9, composite, num_fmt="0.0",
                   fill=score_fill(composite))
        style_cell(ws, row, 10, rec_display,
                   font=Font(bold=True, color=rec_color(rec)),
                   alignment=Alignment(horizontal="center"))
        style_cell(ws, row, 11, signals, font=Font(size=9))

        # Hyperlink to individual sheet
        sheet_name = r["ticker"].replace("-", "_")
        link_cell = style_cell(ws, row, 12, "Go",
                               font=Font(color="0563C1", underline="single"),
                               alignment=Alignment(horizontal="center"))
        link_cell.hyperlink = f"#'{sheet_name}'!A1"

    # Rating legend
    legend_row = len(results_sorted) + 6
    ws.cell(row=legend_row, column=1, value="Rating Scale:").font = Font(bold=True, size=10)
    legends = [
        ("Strong Positive (>= 80)", "C6EFCE"),
        ("Positive (>= 60)", "D6E4F0"),
        ("Neutral (>= 40)", "FFF2CC"),
        ("Negative (< 40)", "FFC7CE"),
    ]
    for j, (text, color) in enumerate(legends):
        cell = ws.cell(row=legend_row, column=3 + j * 2, value=text)
        cell.font = Font(size=9)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = DARK_BLUE


# ============================================================
# Individual stock sheets
# ============================================================

def build_stock_sheet(wb: Workbook, r: dict):
    """Build an individual stock analysis sheet with report + charts."""
    ticker = r["ticker"]
    name = r["name"]
    ohlcv = r["ohlcv"]
    tech = r["tech"]
    fund = r["fund"]
    comp = r["comp"]

    sheet_name = ticker.replace("-", "_")
    ws = wb.create_sheet(title=sheet_name)

    close = ohlcv["Close"]
    current_price = float(close.iloc[-1])
    volume = int(ohlcv["Volume"].iloc[-1])

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 3  # spacer
    for col_letter in "DEFGHIJKLMNOPQRST":
        ws.column_dimensions[col_letter].width = 13

    # ===== REPORT SECTION (Columns A-B, starting row 1) =====
    row = 1

    # Title
    ws.merge_cells(f"A{row}:B{row}")
    title_cell = ws.cell(row=row, column=1, value=f"{name} ({ticker})")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 30
    row += 1

    # Date
    ws.cell(row=row, column=1, value=f"Analysis: {datetime.now().strftime('%Y-%m-%d')}").font = Font(size=9, color=GRAY)
    ws.cell(row=row, column=2, value=f"Data: {ohlcv.index[0].date()} ~ {ohlcv.index[-1].date()}").font = Font(size=9, color=GRAY)
    row += 2

    # --- Overview ---
    def section_header(r, text):
        ws.merge_cells(f"A{r}:B{r}")
        style_cell(ws, r, 1, text, font=SECTION_FONT, fill=SECTION_FILL)
        ws.cell(row=r, column=2).fill = SECTION_FILL
        ws.cell(row=r, column=2).border = THIN
        return r + 1

    def label_value(r, label, value, fmt=None, val_font=None):
        style_cell(ws, r, 1, label, font=LABEL_FONT,
                   fill=PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"))
        f = val_font or VALUE_FONT
        style_cell(ws, r, 2, value, font=f, num_fmt=fmt)
        return r + 1

    row = section_header(row, "Overview")
    market_cap = r.get("market_cap")
    row = label_value(row, "Market Cap", _format_market_cap(market_cap),
                      val_font=Font(bold=True, size=11))
    row = label_value(row, "Current Price", f"${current_price:,.2f}")
    row = label_value(row, "Volume", f"{volume:,}")
    row = label_value(row, "Data Points", f"{len(ohlcv)} days")
    row += 1

    # --- Composite ---
    row = section_header(row, "Composite Evaluation")
    composite = comp["score"]
    rec = comp["recommendation"]
    rec_display = {"strong_positive": "Strong Positive", "positive": "Positive",
                   "neutral": "Neutral", "negative": "Negative"}.get(rec, rec)
    row = label_value(row, "Composite Score", composite, fmt="0.0",
                      val_font=Font(bold=True, size=12, color=rec_color(rec)))
    row = label_value(row, "Recommendation", rec_display,
                      val_font=Font(bold=True, color=rec_color(rec)))
    row = label_value(row, "Tech Weight", "50%")
    row = label_value(row, "Fund Weight", "50%")
    row += 1

    # --- Technical ---
    row = section_header(row, "Technical Analysis")
    row = label_value(row, "Tech Score", tech["score"], fmt="0.0")
    indicators = tech["indicators"]
    row = label_value(row, "RSI (14)", indicators.get("rsi"), fmt="0.00")
    row = label_value(row, "MACD", indicators.get("macd"), fmt="0.0000")
    row = label_value(row, "MACD Signal", indicators.get("macd_signal"), fmt="0.0000")
    row = label_value(row, "MACD Histogram", indicators.get("macd_histogram"), fmt="0.0000")
    row = label_value(row, "SMA 5", indicators.get("sma_5"), fmt="0.00")
    row = label_value(row, "SMA 20", indicators.get("sma_20"), fmt="0.00")
    row = label_value(row, "SMA 60", indicators.get("sma_60"), fmt="0.00")
    row = label_value(row, "SMA 120", indicators.get("sma_120"), fmt="0.00")
    row = label_value(row, "EMA 12", indicators.get("ema_12"), fmt="0.00")
    row = label_value(row, "EMA 26", indicators.get("ema_26"), fmt="0.00")
    row = label_value(row, "BB Upper", indicators.get("bb_upper"), fmt="0.00")
    row = label_value(row, "BB Mid", indicators.get("bb_mid"), fmt="0.00")
    row = label_value(row, "BB Lower", indicators.get("bb_lower"), fmt="0.00")
    row += 1

    # --- Signals ---
    row = section_header(row, "Detected Signals")
    if tech["signals"]:
        for sig in tech["signals"]:
            row = label_value(row, "Signal", sig)
    else:
        row = label_value(row, "Signal", "No notable signals")
    row += 1

    # --- Fundamental ---
    row = section_header(row, "Fundamental Analysis")
    row = label_value(row, "Fund Score", fund["score"], fmt="0.0")
    metrics = fund.get("metrics", {})
    per = metrics.get("per")
    row = label_value(row, "PER", f"{per:.2f}x" if per else "N/A")
    pbr = metrics.get("pbr")
    row = label_value(row, "PBR", f"{pbr:.2f}x" if pbr else "N/A")
    roe = metrics.get("roe")
    row = label_value(row, "ROE", f"{roe*100:.1f}%" if roe and abs(roe) < 10 else "N/A")
    om = metrics.get("operating_margin")
    row = label_value(row, "Op. Margin", f"{om*100:.1f}%" if om and abs(om) < 10 else "N/A")
    rg = metrics.get("revenue_growth")
    row = label_value(row, "Rev. Growth", f"{rg*100:.1f}%" if rg and abs(rg) < 10 else "N/A")
    eg = metrics.get("earnings_growth")
    row = label_value(row, "Earn. Growth", f"{eg*100:.1f}%" if eg and abs(eg) < 10 else "N/A")

    report_end_row = row

    # ===== CHART DATA (Columns D onwards, starting row 1) =====
    # Write chart data headers
    chart_data_start_row = 2
    chart_headers = [
        "Date", "Close", "SMA_20", "SMA_60", "BB_Upper", "BB_Lower",
        "Volume", "RSI_14", "MACD", "MACD_Signal", "MACD_Hist",
    ]
    for j, h in enumerate(chart_headers):
        col = 4 + j  # Column D onwards
        style_cell(ws, chart_data_start_row, col, h, font=HDR_FONT, fill=HDR_FILL,
                   alignment=Alignment(horizontal="center"))

    # Compute indicators for full series
    sma20 = _sma(close, 20)
    sma60 = _sma(close, 60)
    bb = _bbands(close, 20, 2)
    rsi_series = _rsi(close, 14)
    macd_result = _macd(close, 12, 26, 9)

    # Write data rows
    data_rows_count = len(ohlcv)
    for i in range(data_rows_count):
        r_idx = chart_data_start_row + 1 + i
        dt = ohlcv.index[i]
        date_val = dt.tz_localize(None) if dt.tzinfo else dt

        ws.cell(row=r_idx, column=4, value=date_val).number_format = "YYYY-MM-DD"
        ws.cell(row=r_idx, column=5, value=round(float(close.iloc[i]), 2)).number_format = "#,##0.00"

        sma20_val = sma20.iloc[i]
        ws.cell(row=r_idx, column=6, value=round(float(sma20_val), 2) if pd.notna(sma20_val) else None)
        sma60_val = sma60.iloc[i]
        ws.cell(row=r_idx, column=7, value=round(float(sma60_val), 2) if pd.notna(sma60_val) else None)

        bb_u = bb["upper"].iloc[i]
        ws.cell(row=r_idx, column=8, value=round(float(bb_u), 2) if pd.notna(bb_u) else None)
        bb_l = bb["lower"].iloc[i]
        ws.cell(row=r_idx, column=9, value=round(float(bb_l), 2) if pd.notna(bb_l) else None)

        ws.cell(row=r_idx, column=10, value=int(ohlcv["Volume"].iloc[i])).number_format = "#,##0"

        rsi_val = rsi_series.iloc[i]
        ws.cell(row=r_idx, column=11, value=round(float(rsi_val), 2) if pd.notna(rsi_val) else None)

        macd_val = macd_result["macd"].iloc[i]
        ws.cell(row=r_idx, column=12, value=round(float(macd_val), 4) if pd.notna(macd_val) else None)
        sig_val = macd_result["signal"].iloc[i]
        ws.cell(row=r_idx, column=13, value=round(float(sig_val), 4) if pd.notna(sig_val) else None)
        hist_val = macd_result["histogram"].iloc[i]
        ws.cell(row=r_idx, column=14, value=round(float(hist_val), 4) if pd.notna(hist_val) else None)

    data_end_row = chart_data_start_row + data_rows_count

    # ===== CHARTS (placed to the right of report, rows 1~) =====

    # --- Chart 1: Price + SMA + Bollinger (large) ---
    price_chart = LineChart()
    price_chart.title = f"{ticker} Price & Moving Averages"
    price_chart.style = 10
    price_chart.width = 32
    price_chart.height = 16
    price_chart.y_axis.title = "Price (USD)"
    price_chart.x_axis.title = "Date"
    price_chart.legend.position = "b"

    dates_ref = Reference(ws, min_col=4, min_row=chart_data_start_row + 1,
                          max_row=data_end_row)

    # Close price
    close_ref = Reference(ws, min_col=5, min_row=chart_data_start_row,
                          max_row=data_end_row)
    price_chart.add_data(close_ref, titles_from_data=True)
    price_chart.set_categories(dates_ref)
    price_chart.series[0].graphicalProperties.line.width = 20000
    price_chart.series[0].graphicalProperties.line.solidFill = "2F5496"

    # SMA 20
    sma20_ref = Reference(ws, min_col=6, min_row=chart_data_start_row,
                          max_row=data_end_row)
    price_chart.add_data(sma20_ref, titles_from_data=True)
    price_chart.series[1].graphicalProperties.line.solidFill = "ED7D31"
    price_chart.series[1].graphicalProperties.line.dashStyle = "dash"

    # SMA 60
    sma60_ref = Reference(ws, min_col=7, min_row=chart_data_start_row,
                          max_row=data_end_row)
    price_chart.add_data(sma60_ref, titles_from_data=True)
    price_chart.series[2].graphicalProperties.line.solidFill = "70AD47"
    price_chart.series[2].graphicalProperties.line.dashStyle = "dash"

    # BB Upper
    bbu_ref = Reference(ws, min_col=8, min_row=chart_data_start_row,
                        max_row=data_end_row)
    price_chart.add_data(bbu_ref, titles_from_data=True)
    price_chart.series[3].graphicalProperties.line.solidFill = "A5A5A5"
    price_chart.series[3].graphicalProperties.line.dashStyle = "dot"
    price_chart.series[3].graphicalProperties.line.width = 10000

    # BB Lower
    bbl_ref = Reference(ws, min_col=9, min_row=chart_data_start_row,
                        max_row=data_end_row)
    price_chart.add_data(bbl_ref, titles_from_data=True)
    price_chart.series[4].graphicalProperties.line.solidFill = "A5A5A5"
    price_chart.series[4].graphicalProperties.line.dashStyle = "dot"
    price_chart.series[4].graphicalProperties.line.width = 10000

    ws.add_chart(price_chart, "D1")  # Placed at column D, row 1 (right of report)

    # --- Chart 2: RSI ---
    rsi_chart = LineChart()
    rsi_chart.title = "RSI (14)"
    rsi_chart.style = 10
    rsi_chart.width = 32
    rsi_chart.height = 10
    rsi_chart.y_axis.title = "RSI"
    rsi_chart.y_axis.scaling.min = 0
    rsi_chart.y_axis.scaling.max = 100
    rsi_chart.legend.position = "b"

    rsi_ref = Reference(ws, min_col=11, min_row=chart_data_start_row,
                        max_row=data_end_row)
    rsi_chart.add_data(rsi_ref, titles_from_data=True)
    rsi_chart.set_categories(dates_ref)
    rsi_chart.series[0].graphicalProperties.line.solidFill = "7030A0"
    rsi_chart.series[0].graphicalProperties.line.width = 18000

    ws.add_chart(rsi_chart, "D18")

    # --- Chart 3: MACD ---
    macd_chart = LineChart()
    macd_chart.title = "MACD (12, 26, 9)"
    macd_chart.style = 10
    macd_chart.width = 32
    macd_chart.height = 10
    macd_chart.y_axis.title = "MACD"
    macd_chart.legend.position = "b"

    macd_line_ref = Reference(ws, min_col=12, min_row=chart_data_start_row,
                              max_row=data_end_row)
    macd_chart.add_data(macd_line_ref, titles_from_data=True)
    macd_chart.set_categories(dates_ref)
    macd_chart.series[0].graphicalProperties.line.solidFill = "2F5496"

    macd_sig_ref = Reference(ws, min_col=13, min_row=chart_data_start_row,
                             max_row=data_end_row)
    macd_chart.add_data(macd_sig_ref, titles_from_data=True)
    macd_chart.series[1].graphicalProperties.line.solidFill = "ED7D31"
    macd_chart.series[1].graphicalProperties.line.dashStyle = "dash"

    # MACD Histogram as bar chart overlay
    hist_bar = BarChart()
    hist_ref = Reference(ws, min_col=14, min_row=chart_data_start_row,
                         max_row=data_end_row)
    hist_bar.add_data(hist_ref, titles_from_data=True)
    hist_bar.series[0].graphicalProperties.solidFill = "A5A5A5"
    hist_bar.y_axis.axId = 200

    macd_chart += hist_bar

    ws.add_chart(macd_chart, "D31")

    # --- Chart 4: Volume ---
    vol_chart = BarChart()
    vol_chart.title = "Volume"
    vol_chart.style = 10
    vol_chart.width = 32
    vol_chart.height = 8
    vol_chart.y_axis.title = "Volume"
    vol_chart.legend = None

    vol_ref = Reference(ws, min_col=10, min_row=chart_data_start_row,
                        max_row=data_end_row)
    vol_chart.add_data(vol_ref, titles_from_data=True)
    vol_chart.set_categories(dates_ref)
    vol_chart.series[0].graphicalProperties.solidFill = "4472C4"

    ws.add_chart(vol_chart, "D43")

    # Tab color based on recommendation
    color_map = {
        "strong_positive": "00B050",
        "positive": "92D050",
        "neutral": "FFC000",
        "negative": "FF0000",
    }
    ws.sheet_properties.tabColor = color_map.get(comp["recommendation"], "808080")

    # Back to Overview link
    back_cell = ws.cell(row=1, column=14, value="<< Overview")
    back_cell.font = Font(color="0563C1", underline="single", size=9)
    back_cell.hyperlink = "#Overview!A1"

    ws.freeze_panes = "A3"


# ============================================================
# Main
# ============================================================

def main():
    print("=" * 60)
    print("  S&P 500 Top 20 - Full Analysis Report")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)
    print()

    # Step 1: Collect & Analyze
    print("Collecting data and running analysis...")
    print("-" * 60)
    results = collect_and_analyze()
    print("-" * 60)
    print(f"Completed: {len(results)}/20 stocks\n")

    if not results:
        print("ERROR: No data collected. Check network connection.")
        sys.exit(1)

    # Step 2: Build Excel
    print("Building Excel workbook...")
    wb = Workbook()

    # Summary sheet
    build_summary_sheet(wb, results)
    print("  [OK] Overview sheet")

    # Individual sheets
    for i, r in enumerate(results):
        build_stock_sheet(wb, r)
        print(f"  [OK] {r['ticker']} sheet ({i+1}/{len(results)})")

    # Save
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_PATH))
    print(f"\n{'=' * 60}")
    print(f"  Excel saved: {OUTPUT_PATH}")
    print(f"  File size: {OUTPUT_PATH.stat().st_size / 1024:.0f} KB")
    print(f"  Sheets: 1 (Overview) + {len(results)} (Individual)")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
