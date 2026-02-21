"""Export NVIDIA analysis to Excel."""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === 1. NVIDIA OHLCV data ===
from src.collectors.market.us_collector import USMarketCollector

collector = USMarketCollector()
ohlcv = collector.collect_stock_ohlcv("NVDA", days=180)
print(f"OHLCV rows: {len(ohlcv)}")
print(f"Date range: {ohlcv.index[0].date()} ~ {ohlcv.index[-1].date()}")

# === 2. Technical analysis ===
from src.analyzers.technical import TechnicalAnalyzer, _sma, _ema, _rsi, _macd, _bbands

ta = TechnicalAnalyzer()
result = ta.analyze("NVDA", ohlcv=ohlcv)
print(f"Technical Score: {result['score']}")
print(f"Signals: {result['signals']}")

# === 3. Fundamental analysis ===
from src.analyzers.fundamental import FundamentalAnalyzer
from src.core.models import Market

fa = FundamentalAnalyzer()
fund_result = fa.analyze("NVDA", market=Market.US)
print(f"Fundamental Score: {fund_result['score']}")

# === 4. Screener composite ===
from src.analyzers.screener import StockScreener

screener = StockScreener()
screen_result = screener.analyze("NVDA", ohlcv=ohlcv, market=Market.US)
print(f"Composite Score: {screen_result['score']}")
print(f"Recommendation: {screen_result['recommendation']}")

# === 5. Build Excel ===
close = ohlcv["Close"]

# Sheet 1: OHLCV + indicators time series
df_chart = ohlcv.copy()
df_chart.index = df_chart.index.tz_localize(None)
df_chart.index.name = "Date"

df_chart["SMA_5"] = _sma(close, 5)
df_chart["SMA_20"] = _sma(close, 20)
df_chart["SMA_60"] = _sma(close, 60)
df_chart["SMA_120"] = _sma(close, 120)
df_chart["EMA_12"] = _ema(close, 12)
df_chart["EMA_26"] = _ema(close, 26)
df_chart["RSI_14"] = _rsi(close, 14)

macd_df = _macd(close, 12, 26, 9)
df_chart["MACD"] = macd_df["macd"]
df_chart["MACD_Signal"] = macd_df["signal"]
df_chart["MACD_Histogram"] = macd_df["histogram"]

bb_df = _bbands(close, 20, 2)
df_chart["BB_Upper"] = bb_df["upper"]
df_chart["BB_Mid"] = bb_df["mid"]
df_chart["BB_Lower"] = bb_df["lower"]

for col in df_chart.select_dtypes(include=[np.floating]).columns:
    df_chart[col] = df_chart[col].round(2)

# Sheet 2: Summary
rows: list[dict] = []
header_rows: list[int] = []


def add_row(item: str, value: object, is_header: bool = False) -> None:
    if is_header:
        header_rows.append(len(rows))
    rows.append({"item": item, "value": value})


add_row("NVIDIA (NVDA) Analysis Report", "", is_header=True)
add_row("", "")
add_row("ticker", "NVDA")
add_row("analysis date", datetime.now().strftime("%Y-%m-%d"))
add_row("data range", f"{ohlcv.index[0].date()} ~ {ohlcv.index[-1].date()}")
add_row("current price (USD)", f"${float(ohlcv['Close'].iloc[-1]):,.2f}")
add_row("volume", f"{int(ohlcv['Volume'].iloc[-1]):,}")
add_row("", "")

add_row("Technical Analysis", "", is_header=True)
add_row("technical score (0-100)", result["score"])
add_row("RSI (14)", result["indicators"].get("rsi"))
add_row("MACD", result["indicators"].get("macd"))
add_row("MACD Signal", result["indicators"].get("macd_signal"))
add_row("MACD Histogram", result["indicators"].get("macd_histogram"))
add_row("SMA 5d", result["indicators"].get("sma_5"))
add_row("SMA 20d", result["indicators"].get("sma_20"))
add_row("SMA 60d", result["indicators"].get("sma_60"))
add_row("SMA 120d", result["indicators"].get("sma_120"))
add_row("EMA 12d", result["indicators"].get("ema_12"))
add_row("EMA 26d", result["indicators"].get("ema_26"))
add_row("Bollinger Upper", result["indicators"].get("bb_upper"))
add_row("Bollinger Mid", result["indicators"].get("bb_mid"))
add_row("Bollinger Lower", result["indicators"].get("bb_lower"))
add_row("", "")

add_row("Detected Signals", "", is_header=True)
if result["signals"]:
    for sig in result["signals"]:
        add_row("  signal", sig)
else:
    add_row("  signal", "No notable signals")
add_row("", "")

add_row("Fundamental Analysis", "", is_header=True)
add_row("fundamental score (0-100)", fund_result["score"])
metrics = fund_result.get("metrics", {})
metric_labels = {
    "per": "PER (x)",
    "pbr": "PBR (x)",
    "roe": "ROE",
    "operating_margin": "Operating Margin",
    "revenue_growth": "Revenue Growth",
    "earnings_growth": "Earnings Growth",
}
for key, label in metric_labels.items():
    val = metrics.get(key)
    if val is not None:
        if key in ("roe", "operating_margin", "revenue_growth", "earnings_growth"):
            display = f"{val*100:.1f}%" if abs(val) < 10 else f"{val:.1f}%"
        else:
            display = round(val, 2)
    else:
        display = "N/A"
    add_row(label, display)
add_row("", "")

add_row("Composite Evaluation", "", is_header=True)
add_row("composite score (0-100)", screen_result["score"])
rec_map = {
    "strong_positive": "Strong Positive",
    "positive": "Positive",
    "neutral": "Neutral",
    "negative": "Negative",
}
add_row("recommendation", rec_map.get(screen_result["recommendation"], screen_result["recommendation"]))
add_row("weights", "Technical 50% + Fundamental 50%")
add_row("", "")

add_row("Rating Scale", "", is_header=True)
add_row("Strong Positive", ">= 80")
add_row("Positive", ">= 60")
add_row("Neutral", ">= 40")
add_row("Negative", "< 40")

df_summary = pd.DataFrame(rows)
df_summary.columns = ["Item", "Value"]

# Write Excel with formatting
output_path = "data/processed/stocks/NVDA_analysis.xlsx"
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df_chart.to_excel(writer, sheet_name="OHLCV_Indicators", index=True)
    df_summary.to_excel(writer, sheet_name="Analysis_Summary", index=False)

    # --- Format Sheet 1 ---
    ws1 = writer.sheets["OHLCV_Indicators"]
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for cell in ws1[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for col_idx, column_cells in enumerate(ws1.columns, 1):
        max_len = max(len(str(c.value or "")) for c in column_cells)
        ws1.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 20)

    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        for cell in row:
            cell.border = thin
            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"
            elif isinstance(cell.value, int):
                cell.number_format = "#,##0"

    ws1.freeze_panes = "B2"

    # --- Format Sheet 2 ---
    ws2 = writer.sheets["Analysis_Summary"]
    section_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    section_font = Font(bold=True, color="FFFFFF", size=12)

    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 35

    for idx in header_rows:
        row_num = idx + 2  # +1 for 0-index, +1 for header row
        for col_idx in range(1, 3):
            cell = ws2.cell(row=row_num, column=col_idx)
            cell.fill = section_fill
            cell.font = section_font

    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, max_col=2):
        for cell in row:
            cell.border = thin

print(f"\n{'='*50}")
print(f"Excel exported: {output_path}")
print(f"{'='*50}")
print(f"Sheet 1 [OHLCV_Indicators]: {len(df_chart)} rows x {len(df_chart.columns)} cols")
print(f"  OHLCV + SMA(5,20,60,120) + EMA(12,26) + RSI(14)")
print(f"  MACD(12,26,9) + Bollinger Bands(20,2)")
print(f"Sheet 2 [Analysis_Summary]: Technical / Fundamental / Composite")
print(f"{'='*50}")
