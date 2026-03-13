"""Generic single-company research Excel exporter.

Usage:
    cd /home/gint_pcd/projects/주식부자프로젝트
    source .venv-wsl/bin/activate
    python scripts/export_company.py PLTR
    python scripts/export_company.py TSLA
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from src.exporters.base import (
    DARK_BLUE,
    GRAY,
    GREEN,
    HDR_FILL,
    HDR_FONT,
    LABEL_FILL,
    LABEL_FONT,
    LINE_WIDTH_HEAVY,
    LINE_WIDTH_MEDIUM,
    MED_BLUE,
    ORANGE,
    RED,
    SECTION_FILL,
    SECTION_FONT,
    THIN_BORDER,
    TITLE_FONT,
    VALUE_FONT,
    apply_axis_labels,
    apply_chart_gridlines,
    apply_x_axis_tick_interval,
    apply_y_axis_number_format,
    apply_y_axis_padding,
    format_market_cap,
    style_cell,
    style_line_series,
    write_label_value,
    write_section_header,
)
from src.exporters.index_detail_builder import build_index_detail_sheet

# ─── Per-ticker config ────────────────────────────────────────────────────────

COMPETITORS: dict[str, list[str]] = {
    "PLTR": ["SNOW", "DDOG", "AI"],
    "TSLA": ["RIVN", "NIO", "GM"],
    "NVDA": ["AMD", "INTC", "QCOM"],
}

# Quarterly US segment breakdown ($M) — source: earnings releases
SEGMENT_DATA: dict[str, dict] = {
    "PLTR": {
        "quarters": ["Q1'24", "Q2'24", "Q3'24", "Q4'24", "Q1'25", "Q2'25", "Q3'25", "Q4'25"],
        "gov":  [335, 371, 408, 425, 373, 425, 486, 570],
        "comm": [214, 262, 284, 284, 255, 307, 397, 507],
    },
}

THESIS: dict[str, dict] = {
    "PLTR": {
        "bull": [
            "Ontology 해자 — 복제 불가능한 의미론적 레이어",
            "AIP Bootcamp 전환율 70~75% — Commercial 급성장 엔진",
            "US Commercial +137% YoY (Q4 2025)",
            "Rule of 40 = 127% — 업계 최상위권",
            "현금 $7.2B / 부채 없음 — 재무 건전성",
            "2026E 가이던스 $7.18B (+61% YoY)",
        ],
        "bear": [
            "P/S 115x / Forward P/E 150x+ — 극단적 밸류에이션",
            "어닝 미스 1회 = 주가 30~50% 폭락 구조",
            "Forward-deployed 엔지니어 — 스케일 한계",
            "AI 지출 둔화 시 Commercial 성장률 급락",
            "DOGE/예산 삭감 → Government 세그먼트 타격",
            "Jefferies 목표가 $70 (Sell) — 밸류에이션 회의론",
        ],
    },
    "TSLA": {
        "bull": ["FSD/Robo 수익화", "Optimus 로봇", "에너지 사업 성장"],
        "bear": ["EV 점유율 하락", "경영진 집중력 분산", "경쟁 심화"],
    },
}

# ─── Helpers ──────────────────────────────────────────────────────────────────

def _date_to_quarter(dt: object) -> str:
    """Convert a date-like object to a quarter label like Q4'25."""
    month_to_q = {3: "Q1", 6: "Q2", 9: "Q3", 12: "Q4"}
    try:
        m = dt.month
        y = str(dt.year)[-2:]
        return f"{month_to_q.get(m, 'Q?')}'{y}"
    except Exception:
        return str(dt)[:7]


def _safe_float(val: object) -> float | None:
    try:
        f = float(val)
        return None if np.isnan(f) else f
    except (TypeError, ValueError):
        return None


def _fmt_pct(val: float | None) -> str:
    if val is None:
        return "N/A"
    return f"{val * 100:.1f}%"


def _find_row(df: pd.DataFrame, keywords: list[str]) -> pd.Series | None:
    """Find first row in DataFrame whose index matches any keyword."""
    for key in keywords:
        for idx in df.index:
            if key.lower() in str(idx).lower():
                return df.loc[idx]
    return None


# ─── Sheet 1: Summary ─────────────────────────────────────────────────────────

def _build_summary_sheet(
    wb: Workbook,
    ticker: str,
    info: dict,
    ohlcv: pd.DataFrame,
) -> None:
    ws = wb.create_sheet(title="Summary")
    ws.sheet_properties.tabColor = "2F5496"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 44

    # Title
    ws.merge_cells("A1:E1")
    name = info.get("longName", ticker)
    title_cell = ws.cell(row=1, column=1, value=f"{name}  ({ticker})")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.cell(row=2, column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(size=9, color=GRAY)
    ws.cell(row=2, column=4,
            value=f"Sector: {info.get('sector', 'N/A')} / {info.get('industry', '')}").font = Font(size=9, color=GRAY)

    # ── Key Metrics (A:B) ──
    row = 4
    row = write_section_header(ws, row, "Key Metrics")

    close_s = ohlcv["Close"]
    close_price = float(close_s.iloc[-1])
    mkt_cap   = _safe_float(info.get("marketCap"))
    revenue   = _safe_float(info.get("totalRevenue"))
    fcf       = _safe_float(info.get("freeCashflow"))
    pe        = _safe_float(info.get("trailingPE"))
    ps        = _safe_float(info.get("priceToSalesTrailing12Months"))
    op_margin = _safe_float(info.get("operatingMargins"))
    rev_growth = _safe_float(info.get("revenueGrowth"))
    employees = info.get("fullTimeEmployees")

    row = write_label_value(ws, row, "Current Price",
                            f"${close_price:,.2f}",
                            val_font=Font(bold=True, size=12, color=DARK_BLUE))
    row = write_label_value(ws, row, "Market Cap", format_market_cap(mkt_cap))
    row = write_label_value(ws, row, "Revenue (TTM)",
                            f"${revenue/1e9:.2f}B" if revenue else "N/A")
    row = write_label_value(ws, row, "Free Cash Flow (TTM)",
                            f"${fcf/1e9:.2f}B" if fcf else "N/A")
    row = write_label_value(ws, row, "P/E (Trailing)",
                            f"{pe:.1f}x" if pe else "N/A")
    row = write_label_value(ws, row, "P/S (Trailing)",
                            f"{ps:.1f}x" if ps else "N/A")
    row = write_label_value(ws, row, "Operating Margin", _fmt_pct(op_margin))
    row = write_label_value(ws, row, "Revenue Growth YoY", _fmt_pct(rev_growth))
    row = write_label_value(ws, row, "Employees",
                            f"{employees:,}" if employees else "N/A")
    row += 1

    # ── Price Performance (A:B) ──
    row = write_section_header(ws, row, "Price Performance")
    changes = [(2, "1D"), (6, "1W"), (22, "1M"), (65, "3M"), (130, "6M")]
    for n, lbl in changes:
        if len(close_s) >= n:
            chg = (close_s.iloc[-1] / close_s.iloc[-n] - 1) * 100
            color = GREEN if chg >= 0 else RED
            row = write_label_value(
                ws, row, f"{lbl} Change",
                f"{chg:+.2f}%",
                val_font=Font(color=color, bold=True, size=10),
            )
    row += 1

    # 52W high / low
    hi52 = _safe_float(info.get("fiftyTwoWeekHigh"))
    lo52 = _safe_float(info.get("fiftyTwoWeekLow"))
    row = write_section_header(ws, row, "52-Week Range")
    row = write_label_value(ws, row, "52W High",
                            f"${hi52:,.2f}" if hi52 else "N/A")
    row = write_label_value(ws, row, "52W Low",
                            f"${lo52:,.2f}" if lo52 else "N/A")
    if hi52 and lo52 and close_price:
        pos = (close_price - lo52) / (hi52 - lo52) * 100 if hi52 != lo52 else 50
        row = write_label_value(ws, row, "Position in Range",
                                f"{pos:.1f}% (0=Low, 100=High)")

    # ── Bull / Bear (D:E) ──
    thesis = THESIS.get(ticker, {"bull": [], "bear": []})

    bull_fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    bear_fill = PatternFill(start_color="9B0000", end_color="9B0000", fill_type="solid")
    bull_row_fill = PatternFill(start_color="EBF5E1", end_color="EBF5E1", fill_type="solid")
    bear_row_fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")

    brow = 4

    # Bull header (spans D:E)
    ws.merge_cells(f"D{brow}:E{brow}")
    bull_hdr = ws.cell(row=brow, column=4, value="▲  Bull Case")
    bull_hdr.font = Font(bold=True, color="FFFFFF", size=11)
    bull_hdr.fill = bull_fill
    bull_hdr.alignment = Alignment(horizontal="center", vertical="center")
    bull_hdr.border = THIN_BORDER
    ws.cell(row=brow, column=5).fill = bull_fill
    ws.row_dimensions[brow].height = 22
    brow += 1

    for item in thesis["bull"]:
        icon_cell = style_cell(ws, brow, 4, "✓",
                               font=Font(color="375623", bold=True, size=11),
                               fill=bull_row_fill)
        icon_cell.alignment = Alignment(horizontal="center")
        txt_cell = ws.cell(row=brow, column=5, value=item)
        txt_cell.font = Font(size=9)
        txt_cell.fill = bull_row_fill
        txt_cell.border = THIN_BORDER
        txt_cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[brow].height = 18
        brow += 1
    brow += 1

    # Bear header
    ws.merge_cells(f"D{brow}:E{brow}")
    bear_hdr = ws.cell(row=brow, column=4, value="▼  Bear Case")
    bear_hdr.font = Font(bold=True, color="FFFFFF", size=11)
    bear_hdr.fill = bear_fill
    bear_hdr.alignment = Alignment(horizontal="center", vertical="center")
    bear_hdr.border = THIN_BORDER
    ws.cell(row=brow, column=5).fill = bear_fill
    ws.row_dimensions[brow].height = 22
    brow += 1

    for item in thesis["bear"]:
        icon_cell = style_cell(ws, brow, 4, "✗",
                               font=Font(color="9B0000", bold=True, size=11),
                               fill=bear_row_fill)
        icon_cell.alignment = Alignment(horizontal="center")
        txt_cell = ws.cell(row=brow, column=5, value=item)
        txt_cell.font = Font(size=9)
        txt_cell.fill = bear_row_fill
        txt_cell.border = THIN_BORDER
        txt_cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[brow].height = 18
        brow += 1

    ws.freeze_panes = "A3"


# ─── Sheet 2: Financials ──────────────────────────────────────────────────────

def _build_financials_sheet(
    wb: Workbook,
    ticker: str,
    quarterly: pd.DataFrame | None,
) -> None:
    ws = wb.create_sheet(title="Financials")
    ws.sheet_properties.tabColor = "548235"

    # Title
    ws.merge_cells("A1:J1")
    ws.cell(row=1, column=1, value=f"{ticker} — Quarterly Financials").font = TITLE_FONT

    # ── Parse yfinance quarterly data ──
    q_labels: list[str] = []
    rev_vals:  list[float | None] = []
    inc_vals:  list[float | None] = []
    op_vals:   list[float | None] = []

    if quarterly is not None and not quarterly.empty:
        # Most recent first → reverse for chronological order
        cols = list(quarterly.columns)[:8]
        cols_chron = list(reversed(cols))
        q_labels = [_date_to_quarter(c) for c in cols_chron]

        rev_series = _find_row(quarterly, ["Total Revenue", "Revenue"])
        inc_series = _find_row(quarterly, ["Net Income"])
        op_series  = _find_row(quarterly, ["Operating Income", "Operating Revenue"])

        for c in cols_chron:
            rev_vals.append(_safe_float(rev_series[c] / 1e6) if rev_series is not None else None)
            inc_vals.append(_safe_float(inc_series[c] / 1e6) if inc_series is not None else None)
            op_vals.append( _safe_float(op_series[c]  / 1e6) if op_series  is not None else None)

    n_q = len(q_labels)
    HDR_ROW = 3

    # ── Quarterly table header ──
    ws.column_dimensions["A"].width = 22
    style_cell(ws, HDR_ROW, 1, "Metric ($M)",
               font=HDR_FONT, fill=HDR_FILL,
               alignment=Alignment(horizontal="center"))
    for j, lbl in enumerate(q_labels):
        col = 2 + j
        ws.column_dimensions[chr(ord("A") + col - 1)].width = 13
        style_cell(ws, HDR_ROW, col, lbl,
                   font=HDR_FONT, fill=HDR_FILL,
                   alignment=Alignment(horizontal="center"))

    def _write_fin_row(row_num: int, label: str, values: list) -> None:
        style_cell(ws, row_num, 1, label, font=LABEL_FONT, fill=LABEL_FILL)
        for j, v in enumerate(values):
            style_cell(ws, row_num, 2 + j, v, num_fmt="#,##0.0", font=VALUE_FONT)

    _write_fin_row(HDR_ROW + 1, "Revenue ($M)", rev_vals)
    _write_fin_row(HDR_ROW + 2, "Net Income ($M)", inc_vals)
    _write_fin_row(HDR_ROW + 3, "Operating Income ($M)", op_vals)

    # YoY growth row
    style_cell(ws, HDR_ROW + 4, 1, "Revenue Growth YoY%", font=LABEL_FONT, fill=LABEL_FILL)
    for j in range(n_q):
        if j >= 4 and rev_vals[j] and rev_vals[j - 4]:
            g = (rev_vals[j] / rev_vals[j - 4] - 1)
            cell = style_cell(ws, HDR_ROW + 4, 2 + j, g, num_fmt="0.0%")
            cell.font = Font(color=GREEN if g >= 0 else RED, size=10)
        else:
            style_cell(ws, HDR_ROW + 4, 2 + j, None)

    # ── Revenue bar chart ──
    if n_q > 0:
        rev_chart = BarChart()
        rev_chart.type = "col"
        rev_chart.title = f"{ticker} — Quarterly Revenue ($M)"
        rev_chart.style = 10
        rev_chart.width = 28
        rev_chart.height = 16
        rev_chart.legend = None
        apply_chart_gridlines(rev_chart)
        apply_y_axis_number_format(rev_chart, "#,##0")
        apply_axis_labels(rev_chart, y_title="Revenue ($M)")

        rev_ref  = Reference(ws, min_col=2, max_col=1 + n_q,
                             min_row=HDR_ROW + 1, max_row=HDR_ROW + 1)
        cats_ref = Reference(ws, min_col=2, max_col=1 + n_q, min_row=HDR_ROW)
        rev_chart.add_data(rev_ref, from_rows=True, titles_from_data=False)
        rev_chart.set_categories(cats_ref)
        rev_chart.series[0].graphicalProperties.solidFill = "2F5496"
        ws.add_chart(rev_chart, "L3")

    # ── Segment stacked bar chart ──
    seg = SEGMENT_DATA.get(ticker)
    SEG_ROW = HDR_ROW + 7  # start below quarterly table

    if seg:
        n_seg = len(seg["quarters"])

        # Segment table header
        ws.merge_cells(f"A{SEG_ROW - 1}:J{SEG_ROW - 1}")
        style_cell(ws, SEG_ROW - 1, 1,
                   "US Segment Revenue: Government vs Commercial ($M)",
                   font=SECTION_FONT, fill=SECTION_FILL)

        style_cell(ws, SEG_ROW, 1, "Segment",
                   font=HDR_FONT, fill=HDR_FILL,
                   alignment=Alignment(horizontal="center"))
        for j, q in enumerate(seg["quarters"]):
            style_cell(ws, SEG_ROW, 2 + j, q,
                       font=HDR_FONT, fill=HDR_FILL,
                       alignment=Alignment(horizontal="center"))

        style_cell(ws, SEG_ROW + 1, 1, "US Government", font=LABEL_FONT, fill=LABEL_FILL)
        style_cell(ws, SEG_ROW + 2, 1, "US Commercial", font=LABEL_FONT, fill=LABEL_FILL)
        for j, (g, c) in enumerate(zip(seg["gov"], seg["comm"])):
            style_cell(ws, SEG_ROW + 1, 2 + j, g, num_fmt="#,##0")
            style_cell(ws, SEG_ROW + 2, 2 + j, c, num_fmt="#,##0")

        # Stacked bar: each series from one row, col A = label
        seg_chart = BarChart()
        seg_chart.type = "col"
        seg_chart.grouping = "stacked"
        seg_chart.overlap = 100
        seg_chart.title = f"{ticker} — US Segment Mix ($M)"
        seg_chart.style = 10
        seg_chart.width = 28
        seg_chart.height = 16
        apply_chart_gridlines(seg_chart)
        apply_y_axis_number_format(seg_chart, "#,##0")
        apply_axis_labels(seg_chart, y_title="Revenue ($M)")

        # Reference includes col A (label) + data cols
        seg_ref = Reference(ws,
                            min_col=1, max_col=1 + n_seg,
                            min_row=SEG_ROW + 1, max_row=SEG_ROW + 2)
        cats_seg = Reference(ws,
                             min_col=2, max_col=1 + n_seg,
                             min_row=SEG_ROW)
        seg_chart.add_data(seg_ref, from_rows=True, titles_from_data=True)
        seg_chart.set_categories(cats_seg)
        seg_chart.series[0].graphicalProperties.solidFill = "2F5496"   # Gov: blue
        seg_chart.series[1].graphicalProperties.solidFill = "ED7D31"   # Comm: orange

        ws.add_chart(seg_chart, "L20")

    ws.freeze_panes = "B4"


# ─── Sheet 4: Valuation ───────────────────────────────────────────────────────

def _build_valuation_sheet(
    wb: Workbook,
    ticker: str,
    info: dict,
    comp_infos: dict[str, dict],
) -> None:
    ws = wb.create_sheet(title="Valuation")
    ws.sheet_properties.tabColor = "7030A0"

    ws.merge_cells("A1:G1")
    ws.cell(row=1, column=1,
            value=f"{ticker} — Valuation & Peer Comparison").font = TITLE_FONT

    ws.column_dimensions["A"].width = 24
    for col_letter in "BCDEFG":
        ws.column_dimensions[col_letter].width = 16

    all_tickers = [ticker] + list(comp_infos.keys())
    all_infos   = {ticker: info, **comp_infos}

    HDR = 3
    style_cell(ws, HDR, 1, "Metric",
               font=HDR_FONT, fill=HDR_FILL,
               alignment=Alignment(horizontal="center"))
    for j, t in enumerate(all_tickers):
        fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid") if j == 0 else HDR_FILL
        style_cell(ws, HDR, 2 + j, t,
                   font=HDR_FONT, fill=fill,
                   alignment=Alignment(horizontal="center"))

    HIGHLIGHT = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    metrics = [
        ("Market Cap",        "marketCap",                        lambda v: format_market_cap(v)),
        ("P/E (Trailing)",    "trailingPE",                       lambda v: f"{v:.1f}x"),
        ("P/S (Trailing)",    "priceToSalesTrailing12Months",     lambda v: f"{v:.1f}x"),
        ("Revenue Growth",    "revenueGrowth",                    lambda v: f"{v*100:.1f}%"),
        ("Gross Margin",      "grossMargins",                     lambda v: f"{v*100:.1f}%"),
        ("Operating Margin",  "operatingMargins",                 lambda v: f"{v*100:.1f}%"),
        ("FCF (TTM)",         "freeCashflow",                     lambda v: format_market_cap(v)),
        ("Total Revenue TTM", "totalRevenue",                     lambda v: format_market_cap(v)),
        ("52W High",          "fiftyTwoWeekHigh",                 lambda v: f"${v:,.2f}"),
        ("52W Low",           "fiftyTwoWeekLow",                  lambda v: f"${v:,.2f}"),
    ]

    for i, (label, key, fmt_fn) in enumerate(metrics):
        r = HDR + 1 + i
        style_cell(ws, r, 1, label, font=LABEL_FONT, fill=LABEL_FILL)
        for j, t in enumerate(all_tickers):
            raw = _safe_float(all_infos[t].get(key)) if key not in ("marketCap", "freeCashflow", "totalRevenue") \
                  else all_infos[t].get(key)
            try:
                display = fmt_fn(raw) if raw is not None else "N/A"
            except Exception:
                display = "N/A"
            fill = HIGHLIGHT if j == 0 else None
            style_cell(ws, r, 2 + j, display,
                       font=Font(size=10, bold=(j == 0)), fill=fill)

    # ── P/S bar chart (horizontal) ──
    ps_start = HDR + len(metrics) + 3
    ws.merge_cells(f"A{ps_start}:G{ps_start}")
    style_cell(ws, ps_start, 1, "P/S Ratio Comparison",
               font=SECTION_FONT, fill=SECTION_FILL)

    style_cell(ws, ps_start + 1, 1, "Ticker", font=HDR_FONT, fill=HDR_FILL)
    style_cell(ws, ps_start + 1, 2, "P/S (Trailing)", font=HDR_FONT, fill=HDR_FILL)
    for j, t in enumerate(all_tickers):
        ps_val = _safe_float(all_infos[t].get("priceToSalesTrailing12Months"))
        style_cell(ws, ps_start + 2 + j, 1, t, font=LABEL_FONT, fill=LABEL_FILL)
        style_cell(ws, ps_start + 2 + j, 2,
                   round(ps_val, 2) if ps_val else None,
                   num_fmt="0.0")

    n_bars = len(all_tickers)
    ps_chart = BarChart()
    ps_chart.type = "bar"   # horizontal bars
    ps_chart.title = "P/S Ratio Comparison"
    ps_chart.style = 10
    ps_chart.width = 20
    ps_chart.height = 14
    ps_chart.legend = None
    apply_chart_gridlines(ps_chart)
    apply_y_axis_number_format(ps_chart, "0.0")

    ps_data = Reference(ws, min_col=2, max_col=2,
                        min_row=ps_start + 2,
                        max_row=ps_start + 1 + n_bars)
    ps_cats = Reference(ws, min_col=1, max_col=1,
                        min_row=ps_start + 2,
                        max_row=ps_start + 1 + n_bars)
    ps_chart.add_data(ps_data)
    ps_chart.set_categories(ps_cats)
    ps_chart.series[0].graphicalProperties.solidFill = "7030A0"

    ws.add_chart(ps_chart, f"C{ps_start + 1}")
    ws.freeze_panes = "B4"


# ─── Sheet 5: Raw Data ────────────────────────────────────────────────────────

def _build_raw_data_sheet(wb: Workbook, ticker: str, ohlcv: pd.DataFrame) -> None:
    from src.analyzers.technical import _bbands, _ema, _macd, _rsi, _sma
    from src.analyzers.trend import _adx

    ws = wb.create_sheet(title="Raw Data")
    ws.sheet_properties.tabColor = GRAY

    close = ohlcv["Close"]
    df = ohlcv.copy()
    if hasattr(df.index, "tz") and df.index.tz is not None:
        df.index = df.index.tz_localize(None)

    df["SMA_20"]      = _sma(close, 20)
    df["SMA_60"]      = _sma(close, 60)
    df["EMA_12"]      = _ema(close, 12)
    df["EMA_26"]      = _ema(close, 26)
    df["RSI_14"]      = _rsi(close, 14)
    macd_df = _macd(close, 12, 26, 9)
    df["MACD"]        = macd_df["macd"]
    df["MACD_Signal"] = macd_df["signal"]
    df["MACD_Hist"]   = macd_df["histogram"]
    bb_df = _bbands(close, 20, 2)
    df["BB_Upper"]    = bb_df["upper"]
    df["BB_Mid"]      = bb_df["mid"]
    df["BB_Lower"]    = bb_df["lower"]
    if all(c in ohlcv.columns for c in ("High", "Low", "Close")):
        adx_df = _adx(ohlcv)
        df["ADX"]  = adx_df["ADX"]
        df["+DI"]  = adx_df["plus_di"]
        df["-DI"]  = adx_df["minus_di"]

    df = df.round(4)
    df.index.name = "Date"
    df_out = df.reset_index()

    from openpyxl.utils.dataframe import dataframe_to_rows
    for r_idx, row in enumerate(dataframe_to_rows(df_out, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = HDR_FONT
                cell.fill = HDR_FILL
                cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    for col_cells in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 18)

    ws.freeze_panes = "A2"


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    ticker = sys.argv[1].upper() if len(sys.argv) > 1 else "PLTR"

    print(f"\n{'='*60}")
    print(f"  {ticker} Research Excel")
    print(f"{'='*60}")

    # 1. OHLCV
    print("  [1/5] Fetching OHLCV (6mo)...")
    tkr = yf.Ticker(ticker)
    ohlcv = tkr.history(period="6mo")
    if ohlcv.empty:
        print("  ERROR: No OHLCV data. Check ticker or network.")
        sys.exit(1)
    print(f"        {len(ohlcv)} rows | {ohlcv.index[0].date()} ~ {ohlcv.index[-1].date()}")

    # 2. Company info
    print("  [2/5] Fetching company info...")
    info = tkr.info or {}
    print(f"        {info.get('longName', ticker)} | {format_market_cap(info.get('marketCap'))}")

    # 3. Quarterly financials
    print("  [3/5] Fetching quarterly financials...")
    try:
        quarterly = tkr.quarterly_income_stmt
        print(f"        {quarterly.shape[1]} quarters available")
    except Exception as e:
        print(f"        WARNING: {e}")
        quarterly = None

    # 4. Competitor info
    comp_tickers = COMPETITORS.get(ticker, [])[:3]
    print(f"  [4/5] Fetching competitors: {comp_tickers}...")
    comp_infos: dict[str, dict] = {}
    for ct in comp_tickers:
        try:
            ci = yf.Ticker(ct).info or {}
            comp_infos[ct] = ci
            print(f"        {ct}: {format_market_cap(ci.get('marketCap'))}")
        except Exception as e:
            print(f"        WARNING {ct}: {e}")
            comp_infos[ct] = {}

    # 5. Build workbook
    print("  [5/5] Building sheets...")
    wb = Workbook()
    wb.remove(wb.active)

    _build_summary_sheet(wb, ticker, info, ohlcv)
    print("        ✓ Summary")

    _build_financials_sheet(wb, ticker, quarterly)
    print("        ✓ Financials")

    build_index_detail_sheet(wb, "Technical",
                             f"{ticker} — Technical Analysis", ohlcv)
    wb["Technical"].sheet_properties.tabColor = "ED7D31"
    print("        ✓ Technical")

    _build_valuation_sheet(wb, ticker, info, comp_infos)
    print("        ✓ Valuation")

    _build_raw_data_sheet(wb, ticker, ohlcv)
    print("        ✓ Raw Data")

    # Save
    out_dir = ROOT / "data" / "processed" / "stocks"
    out_dir.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now().strftime("%Y%m%d")
    out_path = out_dir / f"{ticker}_research_{date_str}.xlsx"
    wb.save(str(out_path))

    print(f"\n  Saved → {out_path}\n")


if __name__ == "__main__":
    main()
